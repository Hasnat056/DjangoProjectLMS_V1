# Person views
import io
import uuid
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, Sum
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator
from django.views import View
from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.urls import reverse
from django.db import transaction

# Import models from all apps
from .models import Person, Admin, Address, Qualification, Salary, Audittrail, Alumni, ChangeRequest
from AcademicStructure.models import Department, Program, Course, Semester, Semesterdetails, Class
from FacultyModule.models import Faculty, Courseallocation, Lecture, Assessment, Attendance
from StudentModule.models import Student, Enrollment, Result, Transcript

# Import views from other apps for coordination (form-based)
from AcademicStructure import views as academic_views
from FacultyModule import views as faculty_views
from StudentModule import views as student_views

# Form imports
from .forms import AdminProfileForm, SalaryForm, AlumniForm
from StudentModule.forms import StudentForm
from FacultyModule.forms import FacultyForm

def is_admin(user):
    """Check if user is admin"""
    return user.groups.filter(name='Admin').exists()


# ===========================================
# MAIN ADMIN DASHBOARD
# ===========================================

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    """Main admin dashboard with comprehensive system overview"""
    # Get comprehensive statistics
    stats = {
        'total_students': Student.objects.count(),
        'total_faculty': Faculty.objects.count(),
        'total_programs': Program.objects.count(),
        'total_courses': Course.objects.count(),
        'total_departments': Department.objects.count(),
        'active_enrollments': Enrollment.objects.filter(status='Active').count(),
        'active_allocations': Courseallocation.objects.filter(status='Ongoing').count(),
        'total_classes': Class.objects.count(),
        'total_semesters': Semester.objects.count(),
    }

    # Recent activities
    recent_enrollments = Enrollment.objects.select_related(
        'studentid__studentid', 'allocationid__coursecode'
    ).order_by('-enrollmentdate')[:5]

    recent_faculty = Faculty.objects.select_related('employeeid').order_by('-joiningdate')[:5]

    # Department-wise statistics
    dept_stats = []
    for dept in Department.objects.all():
        dept_data = {
            'department': dept,
            'faculty_count': Faculty.objects.filter(departmentid=dept).count(),
            'program_count': Program.objects.filter(departmentid=dept).count(),
            'student_count': Student.objects.filter(programid__departmentid=dept).count(),
        }
        dept_stats.append(dept_data)

    # Performance statistics - FIXED field names
    try:
        avg_gpa = Result.objects.aggregate(Avg('coursegpa'))['coursegpa__avg'] or 0.0  # FIXED: coursegpa
        high_performers = Result.objects.filter(coursegpa__gte=3.5).count()  # FIXED: coursegpa
        total_results = Result.objects.count()
        high_performer_percentage = round((high_performers / total_results * 100), 2) if total_results > 0 else 0
    except:
        avg_gpa = 0.0
        high_performers = 0
        high_performer_percentage = 0

    # Quick stats for widgets
    today = timezone.now().date()
    recent_enrollments_count = Enrollment.objects.filter(
        enrollmentdate__date__gte=today - timedelta(days=7)
    ).count()

    current_session = "Spring 2025"  # Make this dynamic
    active_courses = Courseallocation.objects.filter(
        status='Active',
        session=current_session
    ).count()

    recent_lectures = Lecture.objects.filter(
        startingtime__date__gte=today - timedelta(days=7)
    ).count()

    context = {
        'stats': stats,
        'recent_enrollments': recent_enrollments,
        'recent_faculty': recent_faculty,
        'dept_stats': dept_stats,
        'performance': {
            'avg_gpa': round(float(avg_gpa), 2) if avg_gpa else 0.0,
            'high_performers': high_performers,
            'high_performer_percentage': high_performer_percentage,
        },
        'quick_stats': {
            'recent_enrollments': recent_enrollments_count,
            'active_courses': active_courses,
            'recent_lectures': recent_lectures,
        }
    }

    return render(request, 'admin/dashboard.html', context)


# ===========================================
# GLOBAL SEARCH (Form-based)
# ===========================================

@login_required
@user_passes_test(is_admin)
def global_search(request):
    """Global search across all models with form-based results"""
    query = request.GET.get('q', '')
    model_type = request.GET.get('model', 'all')

    results = {
        'students': [],
        'faculty': [],
        'courses': [],
        'programs': [],
        'enrollments': [],
        'allocations': [],
        'query': query,
        'model_type': model_type
    }

    if query and len(query) >= 2:
        if model_type in ['all', 'students']:
            students = Student.objects.select_related('studentid', 'programid').filter(
                Q(studentid__fname__icontains=query) |
                Q(studentid__lname__icontains=query) |
                Q(studentid__institutionalemail__icontains=query) |
                Q(studentid__personid__icontains=query)
            )[:10]

            results['students'] = students

        if model_type in ['all', 'faculty']:
            faculty = Faculty.objects.select_related('employeeid', 'departmentid').filter(
                Q(employeeid__fname__icontains=query) |
                Q(employeeid__lname__icontains=query) |
                Q(employeeid__institutionalemail__icontains=query) |
                Q(employeeid__personid__icontains=query)
            )[:10]

            results['faculty'] = faculty

        if model_type in ['all', 'courses']:
            courses = Course.objects.filter(
                Q(coursename__icontains=query) |
                Q(coursecode__icontains=query)
            )[:10]

            results['courses'] = courses

        if model_type in ['all', 'programs']:
            programs = Program.objects.select_related('departmentid').filter(
                Q(programname__icontains=query) |
                Q(programid__icontains=query)
            )[:10]

            results['programs'] = programs

        if model_type in ['all', 'enrollments']:
            enrollments = Enrollment.objects.select_related(
                'studentid__studentid', 'allocationid__coursecode'
            ).filter(
                Q(studentid__studentid__fname__icontains=query) |
                Q(studentid__studentid__lname__icontains=query) |
                Q(allocationid__coursecode__coursename__icontains=query) |
                Q(allocationid__coursecode__coursecode__icontains=query)
            )[:10]

            results['enrollments'] = enrollments

        if model_type in ['all', 'allocations']:
            allocations = Courseallocation.objects.select_related(
                'teacherid__employeeid', 'coursecode'
            ).filter(
                Q(teacherid__employeeid__fname__icontains=query) |
                Q(teacherid__employeeid__lname__icontains=query) |
                Q(coursecode__coursename__icontains=query) |
                Q(coursecode__coursecode__icontains=query) |
                Q(session__icontains=query)
            )[:10]

            results['allocations'] = allocations

    # Convert QuerySets to JSON-serializable data
    json_results = {
        'students': [{
            'name': f"{s.studentid.fname} {s.studentid.lname}",
            'email': s.studentid.institutionalemail,
            'id': s.studentid.personid,
            'program': s.programid.programname if s.programid else 'No Program',
            'status': s.status
        } for s in results['students']],

        'faculty': [{
            'name': f"{f.employeeid.fname} {f.employeeid.lname}",
            'email': f.employeeid.institutionalemail,
            'id': f.employeeid.personid,
            'department': f.departmentid.departmentname if f.departmentid else 'No Department',
            'designation': f.designation
        } for f in results['faculty']],

        'courses': [{
            'name': c.coursename,
            'code': c.coursecode,
            'credits': c.credithours,
            'description': c.description if hasattr(c, 'description') else ''
        } for c in results['courses']],

        'programs': [{
            'name': p.programname,
            'id': p.programid,
            'department': p.departmentid.departmentname if p.departmentid else 'No Department',
            'semesters': p.totalsemesters
        } for p in results['programs']],

        'enrollments': [{
            'student_name': f"{e.studentid.studentid.fname} {e.studentid.studentid.lname}",
            'course_name': e.allocationid.coursecode.coursename,
            'course_code': e.allocationid.coursecode.coursecode,
            'status': e.status,
            'enrollment_date': e.enrollmentdate.strftime('%Y-%m-%d') if e.enrollmentdate else ''
        } for e in results['enrollments']],

        'allocations': [{
            'teacher_name': f"{a.teacherid.employeeid.fname} {a.teacherid.employeeid.lname}",
            'course_name': a.coursecode.coursename,
            'course_code': a.coursecode.coursecode,
            'session': a.session,
            'status': a.status
        } for a in results['allocations']],

        'query': query
    }

    return JsonResponse(json_results)

# ===========================================
# API ENDPOINTS FOR DASHBOARD
# ===========================================

@login_required
@user_passes_test(is_admin)
def dashboard_stats_api(request):
    """Get comprehensive dashboard statistics via API"""
    try:
        # Basic counts
        total_students = Student.objects.count()
        total_faculty = Faculty.objects.count()
        total_programs = Program.objects.count()
        total_courses = Course.objects.count()
        total_departments = Department.objects.count()
        total_enrollments = Enrollment.objects.count()

        # Student statistics
        students_by_status = list(
            Student.objects.values('status').annotate(count=Count('studentid')).order_by('status'))
        students_by_program = list(
            Student.objects.values('programid__programname').annotate(count=Count('studentid')).order_by('-count')[:10])

        # Faculty statistics
        faculty_by_department = list(
            Faculty.objects.values('departmentid__departmentname').annotate(count=Count('employeeid')).order_by(
                '-count')
        )

        # Enrollment statistics
        active_enrollments = Enrollment.objects.filter(status='Active').count()
        total_enrollments = Enrollment.objects.count()
        enrollments_by_status = list(
            Enrollment.objects.values('status').annotate(count=Count('enrollmentid')).order_by('status'))

        # Course allocation statistics
        active_allocations = Courseallocation.objects.filter(status='Active').count()
        allocations_by_session = list(
            Courseallocation.objects.values('session').annotate(count=Count('allocationid')).order_by('-session')[:5])

        # Performance statistics - FIXED field names
        avg_gpa = Result.objects.aggregate(Avg('coursegpa'))['coursegpa__avg'] or 0.0  # FIXED: coursegpa
        high_performers = Result.objects.filter(coursegpa__gte=3.5).count()  # FIXED: coursegpa
        total_results = Result.objects.count()

        # All departments with their statistics (even if counts are 0)
        all_departments = []
        for dept in Department.objects.all():
            faculty_count = Faculty.objects.filter(departmentid=dept).count()
            student_count = Student.objects.filter(programid__departmentid=dept).count()
            program_count = Program.objects.filter(departmentid=dept).count()

            all_departments.append({
                'department_name': dept.departmentname,
                'faculty_count': faculty_count,
                'student_count': student_count,
                'program_count': program_count
            })

        stats = {
            'overview': {
                'total_students': total_students,
                'total_faculty': total_faculty,
                'total_programs': total_programs,
                'total_courses': total_courses,
                'total_departments': total_departments,
                'active_enrollments': active_enrollments,
                'active_allocations': active_allocations,
            },
            'departments': all_departments,
            'students': {
                'total': total_students,
                'by_status': students_by_status,
                'by_program': students_by_program,
                'active': Student.objects.filter(status='Active').count()
            },
            'faculty': {
                'total': total_faculty,
                'by_department': faculty_by_department
            },
            'enrollments': {
                'total': total_enrollments,
                'active': active_enrollments,
                'by_status': enrollments_by_status
            },
            'allocations': {
                'active': active_allocations,
                'by_session': allocations_by_session
            },
            'performance': {
                'avg_gpa': round(float(avg_gpa), 2) if avg_gpa else 0.0,
                'high_performers': high_performers,
                'total_results': total_results,
                'high_performer_percentage': round((high_performers / total_results * 100),
                                                   2) if total_results > 0 else 0
            }
        }

        return JsonResponse(stats)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
@user_passes_test(is_admin)
def recent_activities_api(request):
    """Get recent activities for dashboard"""
    try:
        activities = []

        # Recent enrollments
        recent_enrollments = Enrollment.objects.select_related(
            'studentid__studentid', 'allocationid__coursecode'
        ).order_by('-enrollmentdate')[:5]

        for enrollment in recent_enrollments:
            activities.append({
                'type': 'enrollment',
                'title': 'New Enrollment',
                'description': f"{enrollment.studentid.studentid.fname} {enrollment.studentid.studentid.lname} enrolled in {enrollment.allocationid.coursecode.coursename}",
                'timestamp': enrollment.enrollmentdate.isoformat(),
                'link': f'/admin/enrollments/{enrollment.enrollmentid}/'
            })

        # Recent faculty additions
        recent_faculty = Faculty.objects.select_related('employeeid').order_by('-joiningdate')[:3]

        for faculty in recent_faculty:
            activities.append({
                'type': 'faculty',
                'title': 'New Faculty',
                'description': f"{faculty.employeeid.fname} {faculty.employeeid.lname} joined as {faculty.designation}",
                'timestamp': faculty.joiningdate.isoformat(),
                'link': f'/admin/faculty/{faculty.employeeid.personid}/'
            })

        # Recent course allocations
        recent_allocations = Courseallocation.objects.select_related(
            'teacherid__employeeid', 'coursecode'
        ).order_by('-allocationid')[:3]

        for allocation in recent_allocations:
            activities.append({
                'type': 'allocation',
                'title': 'Course Allocation',
                'description': f"{allocation.coursecode.coursename} assigned to {allocation.teacherid.employeeid.fname} {allocation.teacherid.employeeid.lname}",
                'timestamp': datetime.now().isoformat(),  # Allocations don't have timestamp, use current
                'link': f'/admin/allocations/{allocation.allocationid}/'
            })

        # Sort activities by timestamp
        activities.sort(key=lambda x: x['timestamp'], reverse=True)

        return JsonResponse({'activities': activities[:10]})

    except Exception as e:
        return JsonResponse({'activities': []})


@login_required
@user_passes_test(is_admin)
def quick_stats_api(request):
    """Get quick statistics for admin widgets"""
    try:
        # Today's activities
        today = timezone.now().date()

        # Recent enrollments (last 7 days)
        recent_enrollments = Enrollment.objects.filter(
            enrollmentdate__date__gte=today - timedelta(days=7)
        ).count()

        # Active courses this session
        current_session = "Spring 2025"  # You might want to make this dynamic
        active_courses = Courseallocation.objects.filter(
            status='Active',
            session=current_session
        ).count()

        # Students with high GPA (>= 3.5) - FIXED field name
        high_performers = Result.objects.filter(coursegpa__gte=3.5).count()  # FIXED: coursegpa

        # Recent lectures (last 7 days)
        recent_lectures = Lecture.objects.filter(
            startingtime__date__gte=today - timedelta(days=7)
        ).count()

        stats = {
            'recent_enrollments': recent_enrollments,
            'active_courses': active_courses,
            'high_performers': high_performers,
            'recent_lectures': recent_lectures
        }

        return JsonResponse(stats)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
@user_passes_test(is_admin)
def admin_profile_api(request):
    """Get admin profile data as JSON"""
    try:
        admin = get_object_or_404(Admin, employeeid__institutionalemail=request.user.username)
        person = admin.employeeid

        profile_data = {
            'name': f"{person.fname} {person.lname}",
            'email': person.institutionalemail,
            'role': 'System Administrator',
            'person_id': person.personid,
            'initials': f"{person.fname[0]}{person.lname[0]}" if person.fname and person.lname else 'A'
        }

        return JsonResponse(profile_data)

    except Admin.DoesNotExist:
        return JsonResponse({
            'name': 'Administrator',
            'email': '',
            'role': 'System Administrator',
            'initials': 'A'
        })


# ===========================================
# CRUD COORDINATION - Delegate to respective apps (form-based)
# ===========================================

# FACULTY CRUD (delegated to FacultyModule - form-based)
@login_required
@user_passes_test(is_admin)
def faculty_list(request):
    """Delegate to FacultyModule views (form-based)"""
    return faculty_views.faculty_list(request)


@login_required
@user_passes_test(is_admin)
def faculty_create(request):
    """Delegate to FacultyModule views (form-based)"""
    return faculty_views.faculty_create(request)


@login_required
@user_passes_test(is_admin)
def faculty_detail(request, faculty_id):
    """Delegate to FacultyModule views"""
    return faculty_views.faculty_detail(request, faculty_id)


@login_required
@user_passes_test(is_admin)
def faculty_update(request, faculty_id):
    """Delegate to FacultyModule views (form-based)"""
    return faculty_views.faculty_update(request, faculty_id)


@login_required
@user_passes_test(is_admin)
def faculty_delete(request, faculty_id):
    """Delegate to FacultyModule views (form-based)"""
    return faculty_views.faculty_delete(request, faculty_id)


# STUDENT CRUD (delegated to StudentModule - form-based)
@login_required
@user_passes_test(is_admin)
def student_list(request):
    """Delegate to StudentModule views (form-based)"""
    return student_views.student_list(request)


@login_required
@user_passes_test(is_admin)
def student_create(request):
    """Delegate to StudentModule views (form-based)"""
    return student_views.student_create(request)


@login_required
@user_passes_test(is_admin)
def student_detail(request, student_id):
    """Delegate to StudentModule views"""
    return student_views.student_detail(request, student_id)


@login_required
@user_passes_test(is_admin)
def student_update(request, student_id):
    """Delegate to StudentModule views (form-based)"""
    return student_views.student_update(request, student_id)


@login_required
@user_passes_test(is_admin)
def student_delete(request, student_id):
    """Delegate to StudentModule views (form-based)"""
    return student_views.student_delete(request, student_id)


# ACADEMIC STRUCTURE CRUD (delegated to AcademicStructure - form-based)
@login_required
@user_passes_test(is_admin)
def department_list(request):
    """Delegate to AcademicStructure views"""
    return academic_views.department_list(request)



@method_decorator([login_required, user_passes_test(lambda u: u.is_staff)], name='dispatch')
class HODManagementView(View):
    """Handles all HOD change operations"""

    def post(self, request):
        """Handle different HOD operations based on action parameter"""
        action = request.POST.get('action')

        if action == 'submit_change':
            return self.submit_hod_change(request)
        elif action == 'apply_change':
            return self.apply_hod_change(request)
        elif action == 'check_status':
            return self.check_hod_status(request)
        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)

    def get(self, request):
        """Handle GET requests for status checks"""
        dept_id = request.GET.get('dept_id')
        if dept_id:
            return self.check_hod_status(request)
        return JsonResponse({'error': 'Department ID required'}, status=400)

    def submit_hod_change(self, request):
        """Submit HOD change request and send confirmation email"""
        try:
            dept_id = request.POST.get('department_id')
            new_hod_id = request.POST.get('new_hod')
            print (new_hod_id)

            department = get_object_or_404(Department, departmentid=dept_id)
            new_hod = get_object_or_404(Faculty, employeeid=new_hod_id)

            # Check if there's already a pending request
            existing_request = ChangeRequest.objects.filter(
                change_type='hod_change',
                department=department,
                status='pending'
            ).first()

            if existing_request:
                messages.error(request, 'There is already a pending HOD change request for this department.')
                return HttpResponseRedirect('/admin/dashboard?section=departments')

            # Create new change request
            with transaction.atomic():
                change_request = ChangeRequest.objects.create(
                    change_type='hod_change',
                    department=department,
                    new_hod=new_hod,
                    requested_by=request.user,
                    confirmation_token=uuid.uuid4()
                )

                # Send confirmation email
                self.send_confirmation_email(change_request)

            messages.success(request,
                             f'Confirmation email sent successfully to {new_hod.employeeid.fname} {new_hod.employeeid.lname}!')
            return HttpResponseRedirect('/admin/dashboard?section=departments')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return HttpResponseRedirect('/admin/dashboard?section=departments')

    def check_hod_status(self, request):
        """Check status of pending HOD change requests"""
        try:
            dept_id = request.POST.get('dept_id') or request.GET.get('dept_id')
            department = get_object_or_404(Department, departmentid=dept_id)

            # Get pending or confirmed requests
            pending_request = ChangeRequest.objects.filter(
                change_type='hod_change',
                department=department,
                status__in=['pending', 'confirmed']
            ).first()

            if pending_request:
                return JsonResponse({
                    'has_pending_change': True,
                    'status': pending_request.status,
                    'new_hod_name': f"{pending_request.new_hod.employeeid.fname} {pending_request.new_hod.employeeid.lname}",
                    'requested_at': pending_request.requested_at.strftime('%Y-%m-%d %H:%M'),
                    'can_apply': pending_request.status == 'confirmed'
                })
            else:
                return JsonResponse({'has_pending_change': False})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    def apply_hod_change(self, request):
        """Apply confirmed HOD change to department"""
        try:
            dept_id = request.POST.get('department_id')
            department = get_object_or_404(Department, departmentid=dept_id)

            # Get confirmed request
            confirmed_request = ChangeRequest.objects.filter(
                change_type='hod_change',
                department=department,
                status='confirmed'
            ).first()

            if not confirmed_request:
                messages.error(request, 'No confirmed HOD change request found.')
                return HttpResponseRedirect('/admin/dashboard?section=departments')

            # Apply the change
            with transaction.atomic():
                old_hod = department.hod
                department.hod = confirmed_request.new_hod
                department.save()

                # Mark request as applied
                confirmed_request.status = 'applied'
                confirmed_request.applied_at = timezone.now()
                confirmed_request.save()

                self.send_hod_appointment_confirmation(confirmed_request)

                # Optional: Send notification to old HOD
                if old_hod:
                    self.send_hod_change_notification(old_hod, department, confirmed_request.new_hod)

            messages.success(request,
                             f'HOD change applied successfully! {confirmed_request.new_hod.employeeid.fname} {confirmed_request.new_hod.employeeid.lname} is now the Head of Department.')
            return HttpResponseRedirect('/admin/dashboard?section=departments')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return HttpResponseRedirect('/admin/dashboard?section=departments')

    def send_confirmation_email(self, change_request):
        """Send confirmation email to new HOD"""
        confirmation_url = self.request.build_absolute_uri(
            reverse('management:confirm_hod_change', kwargs={'token': change_request.confirmation_token})
        )
        # Add this
        subject = f"HOD Appointment Confirmation - {change_request.department.departmentname}"
        message = f"""
        Dear {change_request.new_hod.employeeid.fname} {change_request.new_hod.employeeid.lname},

        You have been selected as the new Head of Department for {change_request.department.departmentname}.

        Please click the link below to confirm your acceptance:
        {confirmation_url}

        If you decline or have any questions, please contact the administration.

        This link will expire in 48 hours.

        Best regards,
        Namal University, Mianwali
        """

        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[change_request.new_hod.employeeid.personalemail],
            fail_silently=False
        )

    def send_hod_change_notification(self, old_hod, department, new_hod):
        """Send notification to outgoing HOD"""
        subject = f"HOD Change Notification - {department.departmentname}"
        message = f"""
        Dear {old_hod.employeeid.fname} {old_hod.employeeid.lname},

        This is to inform you that the Head of Department role for {department.departmentname} 
        has been transferred to {new_hod.employeeid.fname} {new_hod.employeeid.lname}.

        Thank you for your service as HOD.

        Best regards,
        Namal University, Mianwali
        """

        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[old_hod.employeeid.personalemail],
            fail_silently=True  # Don't fail if email doesn't send
        )

    def send_hod_appointment_confirmation(self, old_hod, department, new_hod):
            """Notify outgoing HOD about the change"""
            try:
                print("DEBUG: Starting send_hod_change_notification")

                if not old_hod:
                    print("DEBUG: No previous HOD to notify")
                    return

                subject = f"HOD Change Notification - {department.departmentname}"

                message = f"""
                Dear {old_hod.employeeid.fname} {old_hod.employeeid.lname},

                We would like to inform you that there has been a change in the Head of Department position for {department.departmentname}.

                {new_hod.employeeid.fname} {new_hod.employeeid.lname} has been appointed as the new Head of Department.

                Thank you for your service as the previous Head of Department. Your contributions to {department.departmentname} are greatly appreciated.

                Best regards,
                Academic Administration Team

                ---
                This is an automated notification about the HOD change.
                """

                recipient_email = old_hod.employeeid.institutionalemail
                print(f"DEBUG: Sending change notification to previous HOD: {recipient_email}")

                send_mail(
                    subject=subject,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[recipient_email],
                    fail_silently=False
                )

                print("DEBUG: HOD change notification sent successfully")

            except Exception as e:
                print(f"DEBUG: Error sending HOD change notification: {e}")
                # Don't raise exception


# Separate view for public email confirmation (no login required)
class HODConfirmationView(View):
    """Handle public email confirmation links"""

    def get(self, request, token):
        """Handle confirmation link clicks"""
        try:
            change_request = get_object_or_404(ChangeRequest, confirmation_token=token)

            # Check if request is still pending
            if change_request.status != 'pending':
                return render(request, 'academic/hod_confirmation.html', {
                    'error': 'This confirmation link has already been used or expired.'
                })

            # Check if request is not too old (48 hours)
            if timezone.now() > change_request.requested_at + timezone.timedelta(hours=48):
                change_request.status = 'expired'
                change_request.save()
                return render(request, 'academic/hod_confirmation.html', {
                    'error': 'This confirmation link has expired.'
                })

            return render(request, 'academic/hod_confirmation.html', {
                'change_request': change_request
            })

        except Exception as e:
            return render(request, 'academic/hod_confirmation.html', {
                'error': 'Invalid confirmation link.'
            })

    def post(self, request, token):
        """Handle confirmation form submission"""
        try:
            change_request = get_object_or_404(ChangeRequest, confirmation_token=token)
            action = request.POST.get('action')


            if change_request.status != 'pending':
                return render(request, 'academic/hod_confirmation.html', {
                    'error': 'This request has already been processed.'
                })

            if action == 'confirm':
                change_request.status = 'confirmed'
                change_request.confirmed_at = timezone.now()
                change_request.save()

                return render(request, 'academic/hod_confirmation.html', {
                    'success': 'Thank you for confirming your HOD appointment.'
                })

            elif action == 'decline':
                change_request.status = 'declined'
                change_request.save()

                return render(request, 'academic/hod_confirmation.html', {
                    'success': 'You have declined the HOD appointment.'
                })

            else:
                # ADD THIS: Handle missing or invalid action
                return render(request, 'academic/hod_confirmation.html', {
                    'error': f'Invalid action: {action}. Please try again.'
                })

        except Exception as e:
            return render(request, 'academic/hod_confirmation.html', {
                'error': 'An error occurred processing your request.'
            })

@login_required
@user_passes_test(is_admin)
def department_detail(request, department_id):
    """Delegate to AcademicStructure views"""
    return academic_views.department_detail(request, department_id)


@login_required
@user_passes_test(is_admin)
def semester_list(request):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.semester_list(request)


@login_required
@user_passes_test(is_admin)
def semester_detail(request,semester_id):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.semester_detail_view(request,semester_id)


@login_required
@user_passes_test(is_admin)
def semester_report(request,semester_id):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.semester_performance_report(request,semester_id)

@login_required
@user_passes_test(is_admin)
def program_list(request):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.program_list(request)


@login_required
@user_passes_test(is_admin)
def program_create(request):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.program_create(request)


@login_required
@user_passes_test(is_admin)
def program_detail(request, program_id):
    """Delegate to AcademicStructure views"""
    return academic_views.program_detail(request, program_id)


@login_required
@user_passes_test(is_admin)
def program_update(request, program_id):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.program_update(request, program_id)


@login_required
@user_passes_test(is_admin)
def program_delete(request, program_id):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.program_delete(request, program_id)


@login_required
@user_passes_test(is_admin)
def course_list(request):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.course_list(request)


@login_required
@user_passes_test(is_admin)
def course_create(request):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.course_create(request)


@login_required
@user_passes_test(is_admin)
def course_detail(request, course_code):
    """Delegate to AcademicStructure views"""
    return academic_views.course_detail(request, course_code)


@login_required
@user_passes_test(is_admin)
def course_update(request, course_code):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.course_update(request, course_code)


@login_required
@user_passes_test(is_admin)
def course_delete(request, course_code):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.course_delete(request, course_code)


@login_required
@user_passes_test(is_admin)
def class_list(request):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.class_list(request)


@login_required
@user_passes_test(is_admin)
def class_create(request):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.class_create(request)


@login_required
@user_passes_test(is_admin)
def class_detail(request, class_id):
    """Delegate to AcademicStructure views"""
    return academic_views.class_detail(request, class_id)


@login_required
@user_passes_test(is_admin)
def class_update(request, class_id):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.class_update(request, class_id)


@login_required
@user_passes_test(is_admin)
def class_delete(request, class_id):
    """Delegate to AcademicStructure views (form-based)"""
    return academic_views.class_delete(request, class_id)


# COURSE ALLOCATIONS CRUD (delegated to FacultyModule - form-based)
@login_required
@user_passes_test(is_admin)
def course_allocation_list(request):
    """Delegate to FacultyModule views (form-based)"""
    return faculty_views.course_allocation_list(request)


@login_required
@user_passes_test(is_admin)
def course_allocation_create(request):
    """Delegate to FacultyModule views (form-based)"""
    return faculty_views.course_allocation_create(request)


@login_required
@user_passes_test(is_admin)
def course_allocation_detail(request, allocation_id):
    """Delegate to FacultyModule views"""
    return faculty_views.course_allocation_detail(request, allocation_id)


@login_required
@user_passes_test(is_admin)
def course_allocation_update(request, allocation_id):
    """Delegate to FacultyModule views (form-based)"""
    return faculty_views.course_allocation_update(request, allocation_id)


@login_required
@user_passes_test(is_admin)
def course_allocation_delete(request, allocation_id):
    """Delegate to FacultyModule views (form-based)"""
    return faculty_views.course_allocation_delete(request, allocation_id)


# ENROLLMENTS CRUD (delegated to StudentModule - form-based)
@login_required
@user_passes_test(is_admin)
def enrollment_list(request):
    """Delegate to StudentModule views (form-based)"""
    return student_views.enrollment_list(request)


@login_required
@user_passes_test(is_admin)
def enrollment_create(request):
    """Delegate to StudentModule views (form-based)"""
    return student_views.enrollment_create(request)


@login_required
@user_passes_test(is_admin)
def enrollment_detail(request, enrollment_id):
    """Delegate to StudentModule views"""
    return student_views.enrollment_detail(request, enrollment_id)


@login_required
@user_passes_test(is_admin)
def enrollment_update(request, enrollment_id):
    """Delegate to StudentModule views (form-based)"""
    return student_views.enrollment_update(request, enrollment_id)


@login_required
@user_passes_test(is_admin)
def enrollment_delete(request, enrollment_id):
    """Delegate to StudentModule views (form-based)"""
    return student_views.enrollment_delete(request, enrollment_id)


# ===========================================
# SALARY MANAGEMENT CRUD (Person app - form-based)
# ===========================================

@login_required
@user_passes_test(is_admin)
def salary_list(request):
    """List all salary records with search and filtering"""
    salaries = Salary.objects.select_related('employeeid').all()

    # Search functionality
    search = request.GET.get('search')
    if search:
        salaries = salaries.filter(
            Q(employeeid__fname__icontains=search) |
            Q(employeeid__lname__icontains=search) |
            Q(employeeid__personid__icontains=search)
        )

    # Year filtering
    year = request.GET.get('year')
    if year:
        salaries = salaries.filter(year=year)

    # Month filtering
    month = request.GET.get('month')
    if month:
        salaries = salaries.filter(month=month)

    # Pagination
    paginator = Paginator(salaries.order_by('-year', '-month'), 25)
    page = request.GET.get('page')
    salaries = paginator.get_page(page)

    return render(request, 'admin/salary_list.html', {
        'salaries': salaries
    })


@login_required
@user_passes_test(is_admin)
def salary_create(request):
    """Create new salary record - FIXED"""
    if request.method == 'POST':
        form = SalaryForm(request.POST)
        if form.is_valid():
            try:
                form.save()  # Form handles duplicate validation
                messages.success(request, 'Salary record created successfully')
                return redirect('person:salary_list')
            except Exception as e:
                messages.error(request, f'Error creating salary record: {str(e)}')
    else:
        form = SalaryForm()

    return render(request, 'admin/salary_create.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def salary_detail(request, salary_id):
    """View salary record details"""
    salary = get_object_or_404(Salary, salaryid=salary_id)

    return render(request, 'admin/salary_detail.html', {
        'salary': salary
    })


@login_required
@user_passes_test(is_admin)
def salary_update(request, salary_id):
    """Update salary record"""
    salary = get_object_or_404(Salary, salaryid=salary_id)

    if request.method == 'POST':
        form = SalaryForm(request.POST, instance=salary)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Salary record updated successfully')
                return redirect('person:salary_list')

            except Exception as e:
                messages.error(request, f'Error updating salary record: {str(e)}')
    else:
        form = SalaryForm(instance=salary)

    return render(request, 'admin/salary_edit.html', {'form': form, 'salary': salary})


@login_required
@user_passes_test(is_admin)
def salary_delete(request, salary_id):
    """Delete salary record"""
    salary = get_object_or_404(Salary, salaryid=salary_id)

    if request.method == 'POST':
        try:
            salary.delete()
            messages.success(request, 'Salary record deleted successfully')
            return redirect('person:salary_list')

        except Exception as e:
            messages.error(request, f'Error deleting salary record: {str(e)}')

    return render(request, 'admin/salary_confirm_delete.html', {'salary': salary})


# ===========================================
# ADMIN VIEW-ONLY FUNCTIONS (Custom implementations)
# ===========================================

# LECTURES via Course Allocations (VIEW ONLY)
@login_required
@user_passes_test(is_admin)
def view_lectures_by_allocation(request, allocation_id):
    """Admin can view lectures through course allocations (READ ONLY)"""
    allocation = get_object_or_404(Courseallocation, allocationid=allocation_id)
    lectures = Lecture.objects.filter(allocationid=allocation).order_by('-startingtime')

    # Search functionality
    search = request.GET.get('search')
    if search:
        lectures = lectures.filter(
            Q(topic__icontains=search) |
            Q(lectureid__icontains=search) |
            Q(venue__icontains=search)
        )

    # Pagination
    paginator = Paginator(lectures, 25)
    page = request.GET.get('page')
    lectures = paginator.get_page(page)

    return render(request, 'admin/lectures_view.html', {
        'allocation': allocation,
        'lectures': lectures,
        'can_modify': False  # Admin cannot modify
    })


@login_required
@user_passes_test(is_admin)
def view_all_lectures_by_allocations(request):
    """Admin can view all lectures organized by allocations"""
    allocations = Courseallocation.objects.select_related(
        'coursecode', 'teacherid__employeeid'
    ).prefetch_related('lecture_set').all()

    return render(request, 'admin/all_lectures_view.html', {
        'allocations': allocations,
        'can_modify': False
    })


# ASSESSMENTS via Course Allocations (VIEW ONLY)
@login_required
@user_passes_test(is_admin)
def view_assessments_by_allocation(request, allocation_id):
    """Admin can view assessments through course allocations (READ ONLY)"""
    allocation = get_object_or_404(Courseallocation, allocationid=allocation_id)
    assessments = Assessment.objects.filter(allocationid=allocation).order_by('-assessmentdate')

    # Search functionality
    search = request.GET.get('search')
    if search:
        assessments = assessments.filter(
            Q(assessmentname__icontains=search) |
            Q(assessmenttype__icontains=search)
        )

    # Pagination
    paginator = Paginator(assessments, 25)
    page = request.GET.get('page')
    assessments = paginator.get_page(page)

    return render(request, 'admin/assessments_view.html', {
        'allocation': allocation,
        'assessments': assessments,
        'can_modify': False  # Admin cannot modify
    })


@login_required
@user_passes_test(is_admin)
def view_all_assessments_by_allocations(request):
    """Admin can view all assessments organized by allocations"""
    allocations = Courseallocation.objects.select_related(
        'coursecode', 'teacherid__employeeid'
    ).prefetch_related('assessment_set').all()

    return render(request, 'admin/all_assessments_view.html', {
        'allocations': allocations,
        'can_modify': False
    })


# ATTENDANCE via Enrollments or Students (VIEW ONLY)
@login_required
@user_passes_test(is_admin)
def view_attendance_by_enrollment(request, enrollment_id):
    """Admin can view attendance through enrollments (READ ONLY)"""
    enrollment = get_object_or_404(Enrollment, enrollmentid=enrollment_id)
    # Get lectures for this enrollment's allocation
    lectures = Lecture.objects.filter(allocationid=enrollment.allocationid).order_by('-startingtime')
    # Get attendance for this student in these lectures
    attendance_records = Attendance.objects.filter(
        studentid=enrollment.studentid,
        lectureid__in=lectures
    ).select_related('lectureid')

    # Calculate attendance statistics - MANUAL calculation since no attendancepercentage field
    total_lectures = lectures.count()
    attended_lectures = attendance_records.count()
    attendance_percentage = (attended_lectures / total_lectures * 100) if total_lectures > 0 else 0

    return render(request, 'admin/attendance_view.html', {
        'enrollment': enrollment,
        'lectures': lectures,
        'attendance_records': attendance_records,
        'total_lectures': total_lectures,
        'attended_lectures': attended_lectures,
        'attendance_percentage': round(attendance_percentage, 2),
        'can_modify': False  # Admin cannot modify
    })


@login_required
@user_passes_test(is_admin)
def view_attendance_by_student(request, student_id):
    """Admin can view attendance through student->enrollments (READ ONLY)"""
    student = get_object_or_404(Student, studentid__personid=student_id)
    enrollments = Enrollment.objects.filter(studentid=student).select_related(
        'allocationid__coursecode'
    )

    attendance_data = []
    for enrollment in enrollments:
        lectures = Lecture.objects.filter(allocationid=enrollment.allocationid)
        attendance = Attendance.objects.filter(
            studentid=student,
            lectureid__in=lectures
        ).select_related('lectureid')

        total_lectures = lectures.count()
        attended_lectures = attendance.count()
        attendance_percentage = (attended_lectures / total_lectures * 100) if total_lectures > 0 else 0

        attendance_data.append({
            'enrollment': enrollment,
            'lectures': lectures,
            'attendance_records': attendance,
            'total_lectures': total_lectures,
            'attended_lectures': attended_lectures,
            'attendance_percentage': round(attendance_percentage, 2)
        })

    return render(request, 'admin/student_attendance_view.html', {
        'student': student,
        'attendance_data': attendance_data,
        'can_modify': False  # Admin cannot modify
    })


# STUDENT ENROLLMENTS, RESULTS, TRANSCRIPTS (VIEW ONLY)
@login_required
@user_passes_test(is_admin)
def view_student_enrollments(request, student_id):
    """Admin can view student enrollments"""
    student = get_object_or_404(Student, studentid__personid=student_id)
    enrollments = Enrollment.objects.filter(studentid=student).select_related(
        'allocationid__coursecode',
        'allocationid__teacherid__employeeid'
    ).order_by('-enrollmentdate')

    # Search functionality
    search = request.GET.get('search')
    if search:
        enrollments = enrollments.filter(
            Q(allocationid__coursecode__coursename__icontains=search) |
            Q(allocationid__coursecode__coursecode__icontains=search) |
            Q(allocationid__session__icontains=search)
        )

    # Pagination
    paginator = Paginator(enrollments, 25)
    page = request.GET.get('page')
    enrollments = paginator.get_page(page)

    return render(request, 'admin/student_enrollments_view.html', {
        'student': student,
        'enrollments': enrollments
    })


@login_required
@user_passes_test(is_admin)
def view_student_results(request, student_id):
    """Admin can view student results"""
    student = get_object_or_404(Student, studentid__personid=student_id)
    results = Result.objects.filter(
        enrollmentid__studentid=student
    ).select_related('enrollmentid__allocationid__coursecode').order_by('-enrollmentid__enrollmentdate')

    # Calculate statistics - FIXED field name
    total_courses = results.count()
    avg_gpa = results.aggregate(Avg('coursegpa'))['coursegpa__avg'] or 0.0  # FIXED: coursegpa

    return render(request, 'admin/student_results_view.html', {
        'student': student,
        'results': results,
        'total_courses': total_courses,
        'avg_gpa': round(float(avg_gpa), 2) if avg_gpa else 0.0
    })


@login_required
@user_passes_test(is_admin)
def view_allocation_results(request, allocation_id):
    """Admin can view results through course allocations"""
    allocation = get_object_or_404(Courseallocation, allocationid=allocation_id)
    results = Result.objects.filter(
        enrollmentid__allocationid=allocation
    ).select_related('enrollmentid__studentid__studentid').order_by('-coursegpa')  # FIXED: coursegpa

    # Calculate statistics - FIXED field name
    total_students = results.count()
    avg_gpa = results.aggregate(Avg('coursegpa'))['coursegpa__avg'] or 0.0  # FIXED: coursegpa
    high_performers = results.filter(coursegpa__gte=3.5).count()  # FIXED: coursegpa

    return render(request, 'admin/allocation_results_view.html', {
        'allocation': allocation,
        'results': results,
        'total_students': total_students,
        'avg_gpa': round(float(avg_gpa), 2) if avg_gpa else 0.0,
        'high_performers': high_performers
    })


@login_required
@user_passes_test(is_admin)
def view_student_transcript(request, student_id):
    """Admin can view student transcripts"""
    student = get_object_or_404(Student, studentid__personid=student_id)
    transcripts = Transcript.objects.filter(studentid=student).select_related('semesterid').order_by(
        'semesterid__semesterno')  # FIXED: semesterno

    # Calculate overall statistics
    total_credits = transcripts.aggregate(Sum('totalcredits'))['totalcredits__sum'] or 0
    overall_gpa = transcripts.aggregate(Avg('semestergpa'))['semestergpa__avg'] or 0.0

    return render(request, 'admin/student_transcript_view.html', {
        'student': student,
        'transcripts': transcripts,
        'total_credits': total_credits,
        'overall_gpa': round(float(overall_gpa), 2) if overall_gpa else 0.0
    })


@login_required
@user_passes_test(is_admin)
def view_class_transcripts(request, class_id):
    """Admin can view transcripts by class"""
    class_obj = get_object_or_404(Class, classid=class_id)
    students = Student.objects.filter(classid=class_obj).select_related('studentid')

    transcript_data = []
    for student in students:
        transcripts = Transcript.objects.filter(studentid=student).select_related('semesterid')
        total_credits = transcripts.aggregate(Sum('totalcredits'))['totalcredits__sum'] or 0
        overall_gpa = transcripts.aggregate(Avg('semestergpa'))['semestergpa__avg'] or 0.0

        transcript_data.append({
            'student': student,
            'transcripts': transcripts,
            'total_credits': total_credits,
            'overall_gpa': round(float(overall_gpa), 2) if overall_gpa else 0.0
        })

    return render(request, 'admin/class_transcripts_view.html', {
        'class_obj': class_obj,
        'transcript_data': transcript_data
    })




@login_required
@user_passes_test(is_admin)
def view_department_students(request, department_id):
    """View department -> students"""
    department = get_object_or_404(Department, departmentid=department_id)
    students = Student.objects.filter(
        programid__departmentid=department
    ).select_related('studentid', 'programid', 'classid').order_by('studentid__fname')

    # Search functionality
    search = request.GET.get('search')
    if search:
        students = students.filter(
            Q(studentid__fname__icontains=search) |
            Q(studentid__lname__icontains=search) |
            Q(studentid__personid__icontains=search)
        )

    # Group by program
    programs = {}
    for student in students:
        if student.programid:
            if student.programid not in programs:
                programs[student.programid] = []
            programs[student.programid].append(student)

    # Pagination
    paginator = Paginator(students, 50)
    page = request.GET.get('page')
    students = paginator.get_page(page)

    return render(request, 'admin/department_students_view.html', {
        'department': department,
        'students': students,
        'programs': programs
    })


# ===========================================
# ADMIN PROFILE MANAGEMENT (Form-based)
# ===========================================

@login_required
@user_passes_test(is_admin)
def admin_profile(request):
    """Admin profile view - edit own profile only"""
    try:
        # Get current admin's person record using your Admin model
        admin = get_object_or_404(Admin, employeeid__institutionalemail=request.user.username)
        person = admin.employeeid

        # Get related data
        try:
            address = Address.objects.get(personid=person)
        except Address.DoesNotExist:
            address = None

        qualifications = Qualification.objects.filter(personid=person)

        return render(request, 'admin/profile.html', {
            'person': person,
            'admin': admin,
            'address': address,
            'qualifications': qualifications
        })

    except Admin.DoesNotExist:
        messages.error(request, 'Admin profile not found.')
        return redirect('management:admin_dashboard')



@login_required
@user_passes_test(is_admin)
def admin_profile_update(request):
    """Update admin's own profile - FIXED VERSION"""
    try:
        admin = get_object_or_404(Admin, employeeid__institutionalemail=request.user.username)

        if request.method == 'POST':
            form = AdminProfileForm(request.POST, instance=admin)

            if form.is_valid():
                try:
                    form.save()  # Form now handles everything
                    messages.success(request, 'Profile updated successfully')
                    return redirect('person:admin_profile')
                except Exception as e:
                    messages.error(request, f'Error updating profile: {str(e)}')
        else:
            form = AdminProfileForm(instance=admin)

        return render(request, 'admin/profile_edit.html', {'form': form, 'admin': admin})

    except Admin.DoesNotExist:
        messages.error(request, 'Admin profile not found.')
        return redirect('/admin/dashboard/?section=admin/')


# ===========================================
# ALUMNI MANAGEMENT CRUD (Person app - form-based)
# ===========================================

@login_required
@user_passes_test(is_admin)
def alumni_list(request):
    """List all alumni records with search and filtering"""
    alumni = Alumni.objects.select_related(
        'alumniid__studentid',
        'alumniid__programid__departmentid'
    ).all()

    # Search functionality
    search = request.GET.get('search')
    if search:
        alumni = alumni.filter(
            Q(alumniid__studentid__fname__icontains=search) |
            Q(alumniid__studentid__lname__icontains=search) |
            Q(alumniid__studentid__personid__icontains=search) |
            Q(email__icontains=search)
        )

    # Filter by graduation year
    year = request.GET.get('year')
    if year:
        alumni = alumni.filter(graduationdate__year=year)

    # Filter by program
    program = request.GET.get('program')
    if program:
        alumni = alumni.filter(alumniid__programid=program)

    # Pagination
    paginator = Paginator(alumni.order_by('-graduationdate'), 25)
    page = request.GET.get('page')
    alumni = paginator.get_page(page)

    # Get programs for filter dropdown
    programs = Program.objects.all()

    return render(request, 'admin/alumni_list.html', {
        'alumni': alumni,
        'programs': programs
    })


@login_required
@user_passes_test(is_admin)
def alumni_create(request):
    """Create new alumni record"""
    if request.method == 'POST':
        form = AlumniForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Alumni record created successfully')
                return redirect('person:alumni_list')
            except Exception as e:
                messages.error(request, f'Error creating alumni record: {str(e)}')
    else:
        form = AlumniForm()

    return render(request, 'admin/alumni_create.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def alumni_detail(request, alumni_id):
    """View alumni record details"""
    # Note: alumni_id is the student's personid
    alumni = get_object_or_404(Alumni, alumniid__studentid__personid=alumni_id)

    # Get related information
    student = alumni.alumniid
    person = student.studentid

    # Get academic records
    enrollments = Enrollment.objects.filter(studentid=student).select_related(
        'allocationid__coursecode'
    )
    results = Result.objects.filter(enrollmentid__in=enrollments).select_related(
        'enrollmentid__allocationid__coursecode'
    )

    # Calculate final GPA
    avg_gpa = results.aggregate(Avg('coursegpa'))['coursegpa__avg'] or 0.0

    return render(request, 'admin/alumni_detail.html', {
        'alumni': alumni,
        'student': student,
        'person': person,
        'enrollments': enrollments,
        'results': results,
        'avg_gpa': round(float(avg_gpa), 2) if avg_gpa else 0.0
    })


@login_required
@user_passes_test(is_admin)
def alumni_update(request, alumni_id):
    """Update alumni record"""
    alumni = get_object_or_404(Alumni, alumniid__studentid__personid=alumni_id)

    if request.method == 'POST':
        form = AlumniForm(request.POST, instance=alumni)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Alumni record updated successfully')
                return redirect('person:alumni_list')
            except Exception as e:
                messages.error(request, f'Error updating alumni record: {str(e)}')
    else:
        form = AlumniForm(instance=alumni)

    return render(request, 'admin/alumni_edit.html', {
        'form': form,
        'alumni': alumni
    })


@login_required
@user_passes_test(is_admin)
def alumni_delete(request, alumni_id):
    """Delete alumni record"""
    alumni = get_object_or_404(Alumni, alumniid__studentid__personid=alumni_id)

    if request.method == 'POST':
        try:
            alumni.delete()
            messages.success(request, 'Alumni record deleted successfully')
            return redirect('person:alumni_list')
        except Exception as e:
            messages.error(request, f'Error deleting alumni record: {str(e)}')

    return render(request, 'admin/alumni_confirm_delete.html', {'alumni': alumni})



# ===========================================
# REPORTING FUNCTIONS (Form-based)
# ===========================================

@login_required
@user_passes_test(is_admin)
def reports(request):
    total_students = Student.objects.all().count()
    enrolled_students = Student.objects.all().filter(status='Enrolled').count()
    graduated_students = Student.objects.all().filter(status='Graduated').count()
    dropped_students = Student.objects.all().filter(status='Dropped').count()

    total_faculty = Faculty.objects.all().count()
    total_enrollments = Enrollment.objects.all().count()
    active_enrollments = Enrollment.objects.all().filter(status='Active').count()
    total_allocations = Courseallocation.objects.all().count()
    ongoing_allocations = Courseallocation.objects.all().filter(status='Ongoing').count()
    total_semesters_held = Semester.objects.all().count()
    total_programs = Program.objects.all().count()
    total_courses = Course.objects.all().count()
    active_courses = Course.objects.all().filter(courseallocation__status='Ongoing').distinct().count()
    total_classes = Class.objects.all().count()

    department_based_programs = Department.objects.annotate(total=Count('program')).values('departmentname', 'total')
    department_based_faculty = Department.objects.annotate(total=Count('faculty')).values('departmentname', 'total')
    department_based_classes = Department.objects.annotate(total=Count('program__class')).values('departmentname', 'total')

    if total_students > 0:
        graduation_rate = round((graduated_students / total_students) * 100, 1)
    else:
        graduation_rate = 0

    if total_students > 0:
        dropped_rate = round((dropped_students / total_students) * 100, 1)
    else:
        dropped_rate = 0

    if total_courses > 0:
        course_activity_rate = round((active_courses / total_courses) * 100, 1)
    else:
        course_activity_rate = 0

    if total_enrollments > 0:
        enrollment_activity_rate = round((active_enrollments / total_enrollments) * 100, 1)
    else:
        enrollment_activity_rate = 0

    if total_allocations > 0:
        allocation_activity_rate = round ((ongoing_allocations / total_allocations) * 100, 1)
    else:
        allocation_activity_rate = 0



    context = {
        'total_students': total_students,
        'enrolled_students': enrolled_students,
        'graduated_students': graduated_students,
        'dropped_students': dropped_students,
        'total_faculty': total_faculty,
        'total_enrollments': total_enrollments,
        'active_enrollments': active_enrollments,
        'total_allocations': total_allocations,
        'ongoing_allocations': ongoing_allocations,
        'total_semesters_held': total_semesters_held,
        'total_programs': total_programs,
        'total_courses': total_courses,
        'active_courses': active_courses,
        'total_classes': total_classes,
        'department_based_faculty': department_based_faculty,
        'department_based_programs': department_based_programs,
        'department_based_classes': department_based_classes,
        'graduation_rate': graduation_rate,
        'dropped_rate': dropped_rate,
        'course_activity_rate': course_activity_rate,
        'enrollment_activity_rate': enrollment_activity_rate,
        'allocation_activity_rate': allocation_activity_rate,

    }

    #print(context)

    return render (request,'admin/reports.html', context)




@login_required
@user_passes_test(is_admin)
def alumni_report(request):
    """Generate alumni report with statistics"""
    alumni = Alumni.objects.select_related(
        'alumniid__studentid',
        'alumniid__programid__departmentid'
    ).all()

    # Alumni by year
    alumni_by_year = alumni.extra(
        select={'year': "EXTRACT(year FROM graduationdate)"}
    ).values('year').annotate(count=Count('alumniid')).order_by('-year')

    # Alumni by program
    alumni_by_program = alumni.values(
        'alumniid__programid__programname'
    ).annotate(count=Count('alumniid')).order_by('-count')

    # Alumni by department
    alumni_by_department = alumni.values(
        'alumniid__programid__departmentid__departmentname'
    ).annotate(count=Count('alumniid')).order_by('-count')

    # Employment statistics
    employed_alumni = alumni.exclude(employmentinfo='').exclude(employmentinfo__isnull=True).count()
    total_alumni = alumni.count()
    employment_rate = (employed_alumni / total_alumni * 100) if total_alumni > 0 else 0

    context = {
        'total_alumni': total_alumni,
        'employed_alumni': employed_alumni,
        'employment_rate': round(employment_rate, 2),
        'alumni_by_year': alumni_by_year,
        'alumni_by_program': alumni_by_program,
        'alumni_by_department': alumni_by_department,
    }

    #print(context)

    return render(request, 'admin/alumni_report.html', context)



@login_required
@user_passes_test(is_admin)
def generate_department_report(request, department_id):
    """Generate comprehensive department report"""
    department = get_object_or_404(Department, departmentid=department_id)

    # Department statistics
    programs = Program.objects.filter(departmentid=department)
    faculty = Faculty.objects.filter(departmentid=department)
    students = Student.objects.filter(programid__departmentid=department)

    # Course allocations in this department
    allocations = Courseallocation.objects.filter(
        coursecode__semesterdetails__semesterid__programid__departmentid=department  # FIXED: programid path
    ).select_related('coursecode', 'teacherid__employeeid')

    # Student performance by program - FIXED field name
    program_performance = []
    for program in programs:
        program_students = students.filter(programid=program)

        # Get average GPA for this program (if available)
        try:
            avg_gpa = Result.objects.filter(
                enrollmentid__studentid__programid=program
            ).aggregate(avg_gpa=Avg('coursegpa'))['coursegpa__avg']  # FIXED: coursegpa
        except:
            avg_gpa = None

        program_performance.append({
            'program': program,
            'student_count': program_students.count(),
            'avg_gpa': round(avg_gpa, 2) if avg_gpa else 'N/A'
        })

    # Faculty workload
    faculty_workload = []
    for f in faculty:
        f_allocations = allocations.filter(teacherid=f)

        faculty_workload.append({
            'faculty': f,
            'allocation_count': f_allocations.count(),
            'courses': [alloc.coursecode for alloc in f_allocations]
        })

    context = {
        'department': department,
        'programs': programs,
        'faculty': faculty,
        'students': students,
        'allocations': allocations,
        'program_performance': program_performance,
        'faculty_workload': faculty_workload,
        'stats': {
            'total_programs': programs.count(),
            'total_faculty': faculty.count(),
            'total_students': students.count(),
            'total_allocations': allocations.count(),
        }
    }
    #print(context)

    return render(request, 'admin/department_report.html', context)


def faculty_performance_report(request):
    """Enhanced faculty performance report with complete context data"""
    # Fix: Remove the problematic select_related for personid
    faculty = Faculty.objects.select_related('employeeid', 'departmentid').all()

    faculty_performance = []
    total_students_all = 0
    total_allocations_all = 0
    gpa_values = []

    for f in faculty:
        # Get allocations for this faculty
        allocations = Courseallocation.objects.filter(teacherid=f).select_related('coursecode')

        # Get enrollments for faculty's courses
        enrollments = Enrollment.objects.filter(allocationid__in=allocations)

        # Calculate performance metrics
        total_students = enrollments.count()
        total_students_all += total_students
        total_allocations_all += allocations.count()

        # Average student performance in faculty's courses
        try:
            results = Result.objects.filter(enrollmentid__in=enrollments)
            if results.exists():
                avg_student_gpa = results.aggregate(avg_gpa=Avg('coursegpa'))['avg_gpa']
                avg_student_gpa = round(avg_student_gpa, 2) if avg_student_gpa else None
                if avg_student_gpa:
                    gpa_values.append(avg_student_gpa)
            else:
                avg_student_gpa = None
        except:
            avg_student_gpa = None

        # Determine performance level
        if avg_student_gpa:
            if avg_student_gpa >= 3.5:
                performance_level = "Excellent"
            elif avg_student_gpa >= 3.0:
                performance_level = "Good"
            elif avg_student_gpa >= 2.5:
                performance_level = "Average"
            else:
                performance_level = "Needs Improvement"
        else:
            performance_level = "No Data"

        # Determine workload level
        allocation_count = allocations.count()
        if allocation_count >= 4:
            workload_level = "High"
        elif allocation_count >= 2:
            workload_level = "Medium"
        else:
            workload_level = "Light"

        # Access faculty name properly based on your model structure
        try:
            # Try to get name from employeeid if it has name fields
            if hasattr(f.employeeid, 'fname') and hasattr(f.employeeid, 'lname'):
                faculty_name = f"{f.employeeid.fname} {f.employeeid.lname}"
                faculty_id = str(f.employeeid.pk)  # Use primary key as ID
                # Generate initials safely
                fname = f.employeeid.fname.strip() if f.employeeid.fname else "F"
                lname = f.employeeid.lname.strip() if f.employeeid.lname else "L"
                faculty_initials = f"{fname[0]}{lname[0]}"
            else:
                # Fallback to string representation or employee ID
                faculty_name = str(f.employeeid)
                faculty_id = str(f.employeeid.pk)
                faculty_initials = faculty_name[:2].upper() if len(faculty_name) >= 2 else "FA"
        except:
            faculty_name = f"Faculty {f.pk}"
            faculty_id = str(f.pk)
            faculty_initials = "FA"

        # Get department name safely
        try:
            department_name = f.departmentid.departmentname if f.departmentid else "Unknown Department"
        except:
            department_name = "Unknown Department"

        faculty_performance.append({
            'faculty': f,
            'faculty_name': faculty_name,
            'faculty_id': faculty_id,
            'faculty_initials': faculty_initials,  # Add initials
            'department_name': department_name,
            'allocation_count': allocation_count,
            'allocations': allocations,
            'enrollments': enrollments,
            'total_students': total_students,
            'avg_student_gpa': avg_student_gpa,
            'performance_level': performance_level,
            'workload_level': workload_level,
            'course_codes': [alloc.coursecode.coursecode for alloc in allocations] if allocations else []
        })

    # Calculate overall statistics
    overall_avg_gpa = sum(gpa_values) / len(gpa_values) if gpa_values else None

    # Performance distribution
    performance_stats = {
        'excellent': len([f for f in faculty_performance if f['performance_level'] == 'Excellent']),
        'good': len([f for f in faculty_performance if f['performance_level'] == 'Good']),
        'average': len([f for f in faculty_performance if f['performance_level'] == 'Average']),
        'needs_improvement': len([f for f in faculty_performance if f['performance_level'] == 'Needs Improvement']),
        'no_data': len([f for f in faculty_performance if f['performance_level'] == 'No Data'])
    }

    # Workload distribution
    workload_stats = {
        'high': len([f for f in faculty_performance if f['workload_level'] == 'High']),
        'medium': len([f for f in faculty_performance if f['workload_level'] == 'Medium']),
        'light': len([f for f in faculty_performance if f['workload_level'] == 'Light'])
    }

    # Department-wise statistics
    departments = {}
    for f in faculty_performance:
        dept_name = f['department_name']
        if dept_name not in departments:
            departments[dept_name] = {
                'faculty_count': 0,
                'total_students': 0,
                'total_allocations': 0,
                'avg_gpas': []
            }
        departments[dept_name]['faculty_count'] += 1
        departments[dept_name]['total_students'] += f['total_students']
        departments[dept_name]['total_allocations'] += f['allocation_count']
        if f['avg_student_gpa']:
            departments[dept_name]['avg_gpas'].append(f['avg_student_gpa'])

    # Calculate department averages
    for dept in departments.values():
        dept['avg_gpa'] = sum(dept['avg_gpas']) / len(dept['avg_gpas']) if dept['avg_gpas'] else None

    # Sort by performance metrics
    sort_by = request.GET.get('sort', 'avg_student_gpa')
    reverse = request.GET.get('order', 'desc') == 'desc'

    if sort_by == 'avg_student_gpa':
        faculty_performance.sort(
            key=lambda x: x['avg_student_gpa'] if x['avg_student_gpa'] else -1,
            reverse=reverse
        )
    elif sort_by == 'allocation_count':
        faculty_performance.sort(key=lambda x: x['allocation_count'], reverse=reverse)
    elif sort_by == 'total_students':
        faculty_performance.sort(key=lambda x: x['total_students'], reverse=reverse)
    elif sort_by == 'faculty_name':
        faculty_performance.sort(key=lambda x: x['faculty_name'], reverse=reverse)

    # Generate insights and recommendations
    insights = generate_faculty_insights(faculty_performance, performance_stats, workload_stats)

    context = {
        'faculty_performance': faculty_performance,
        'sort_by': sort_by,
        'order': 'desc' if reverse else 'asc',

        # Overall statistics
        'total_faculty': len(faculty_performance),
        'total_students': total_students_all,
        'total_allocations': total_allocations_all,
        'overall_avg_gpa': round(overall_avg_gpa, 2) if overall_avg_gpa else None,

        # Performance distribution
        'performance_stats': performance_stats,
        'workload_stats': workload_stats,

        # Department statistics
        'departments': departments,

        # Insights and recommendations
        'insights': insights,

        # Additional metrics
        'faculty_utilization_rate': round((total_allocations_all / len(faculty_performance)) * 100 / 4,
                                          1) if faculty_performance else 0,  # Assuming 4 is max courses
        'avg_students_per_faculty': round(total_students_all / len(faculty_performance),
                                          1) if faculty_performance else 0,
    }

    return render(request, 'admin/faculty_performance_report.html', context)


def generate_faculty_insights(faculty_performance, performance_stats, workload_stats):
    """Generate dynamic insights and recommendations based on faculty data"""
    insights = {
        'recommendations': [],
        'action_items': [],
        'key_findings': []
    }

    total_faculty = len(faculty_performance)

    # Performance-based recommendations
    if performance_stats['needs_improvement'] > 0:
        insights['recommendations'].append(
            f"Provide additional support for {performance_stats['needs_improvement']} faculty with low student performance"
        )
        insights['action_items'].append({
            'priority': 'high',
            'text': f'Schedule performance review for {performance_stats["needs_improvement"]} faculty members'
        })

    if performance_stats['excellent'] > 0:
        insights['recommendations'].append(
            f"Recognize and share best practices from {performance_stats['excellent']} high-performing faculty"
        )
        insights['action_items'].append({
            'priority': 'medium',
            'text': 'Organize best practices sharing session'
        })

    # Workload-based recommendations
    if workload_stats['high'] > workload_stats['medium']:
        insights['recommendations'].append(
            "Consider redistributing course load to balance faculty workload"
        )
        insights['action_items'].append({
            'priority': 'high',
            'text': 'Review and rebalance course allocations'
        })

    if workload_stats['light'] > 0:
        insights['recommendations'].append(
            f"Optimize course allocation for {workload_stats['light']} faculty with light teaching loads"
        )

    # Key findings
    if performance_stats['no_data'] > total_faculty * 0.3:
        insights['key_findings'].append(
            f"{performance_stats['no_data']} faculty members have no performance data available"
        )

    if workload_stats['high'] > 0:
        insights['key_findings'].append(
            f"{workload_stats['high']} faculty members have high course loads (4+ courses)"
        )

    # Default action items if none generated
    if not insights['action_items']:
        insights['action_items'] = [
            {'priority': 'medium', 'text': 'Continue monitoring faculty performance'},
            {'priority': 'low', 'text': 'Plan quarterly faculty review meetings'}
        ]

    return insights


from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Avg, Q, F, Max, Min, Case, When, DecimalField
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal


@login_required
@user_passes_test(is_admin)
def student_analytics_report(request):
    """Enhanced comprehensive student analytics report"""

    # =============================================================================
    # 1. BASIC STUDENT STATISTICS
    # =============================================================================
    total_students = Student.objects.count()
    active_students = Student.objects.filter(status='Enrolled').count()
    graduated_students = Student.objects.filter(status='Graduated').count()
    dropped_students = Student.objects.filter(status='Dropped').count()

    # =============================================================================
    # 2. BATCH-BASED ANALYTICS (Your Primary Grouping)
    # =============================================================================
    def get_batch_analytics():
        batch_stats = Student.objects.values(
            'classid__batchyear',
            'classid__programid__programname',
            'classid__programid',
            'classid__programid__departmentid__departmentname'
        ).annotate(
            total_students=Count('studentid'),
            enrolled_students=Count('studentid', filter=Q(status='Enrolled')),
            graduated_students=Count('studentid', filter=Q(status='Graduated')),
            dropped_students=Count('studentid', filter=Q(status='Dropped')),
            # Calculate average GPA for the batch
            avg_gpa=Avg('enrollment__result__coursegpa'),
            # Total enrollments by batch
            total_enrollments=Count('enrollment'),
            # Completed courses
            completed_courses=Count('enrollment', filter=Q(enrollment__status='Completed'))
        ).order_by('-classid__batchyear')

        # Calculate additional metrics for each batch
        for batch in batch_stats:
            if batch['total_students'] > 0:
                batch['graduation_rate'] = (batch['graduated_students'] / batch['total_students']) * 100
                batch['dropout_rate'] = (batch['dropped_students'] / batch['total_students']) * 100
                batch['retention_rate'] = (batch['enrolled_students'] / batch['total_students']) * 100
            else:
                batch['graduation_rate'] = batch['dropout_rate'] = batch['retention_rate'] = 0

        return batch_stats

    # =============================================================================
    # 3. PROGRAM & DEPARTMENT ANALYTICS
    # =============================================================================
    def get_program_analytics():
        return Student.objects.values(
            'programid__programname',
            'programid__departmentid__departmentname',
            'programid__programid'
        ).annotate(
            student_count=Count('studentid'),
            enrolled_count=Count('studentid', filter=Q(status='Enrolled')),
            graduated_count=Count('studentid', filter=Q(status='Graduated')),
            dropped_count=Count('studentid', filter=Q(status='Dropped')),
            avg_program_gpa=Avg('enrollment__result__coursegpa'),
            total_credit_hours=Count('enrollment__result', filter=Q(enrollment__result__coursegpa__gte=2.0)),
            avg_enrollments_per_student=Avg('enrollment__enrollmentid')
        ).order_by('-student_count')

    def get_department_analytics():
        return Student.objects.values(
            'programid__departmentid__departmentname'
        ).annotate(
            student_count=Count('studentid'),
            program_count=Count('programid', distinct=True),
            avg_dept_gpa=Avg('enrollment__result__coursegpa'),
            enrolled_students=Count('studentid', filter=Q(status='Enrolled')),
            graduation_rate=Count('studentid', filter=Q(status='Graduated')) * 100.0 / Count('studentid')
        ).order_by('-student_count')

    # =============================================================================
    # 4. ACADEMIC PERFORMANCE ANALYTICS
    # =============================================================================
    def get_performance_analytics():
        # Overall GPA statistics
        gpa_stats = Result.objects.aggregate(
            overall_avg_gpa=Avg('coursegpa'),
            highest_gpa=Max('coursegpa'),
            lowest_gpa=Min('coursegpa')
        )

        # GPA distribution
        gpa_distribution = Result.objects.aggregate(
            excellent_count=Count('resultid', filter=Q(coursegpa__gte=3.5)),
            good_count=Count('resultid', filter=Q(coursegpa__gte=3.0, coursegpa__lt=3.5)),
            satisfactory_count=Count('resultid', filter=Q(coursegpa__gte=2.5, coursegpa__lt=3.0)),
            poor_count=Count('resultid', filter=Q(coursegpa__lt=2.5))
        )

        # Top performing students
        top_students = Student.objects.annotate(
            student_avg_gpa=Avg('enrollment__result__coursegpa'),
            total_courses=Count('enrollment__result')
        ).filter(
            student_avg_gpa__isnull=False,
            total_courses__gte=3  # At least 3 completed courses
        ).order_by('-student_avg_gpa')[:10]

        return {
            'gpa_stats': gpa_stats,
            'gpa_distribution': gpa_distribution,
            'top_students': top_students
        }

    # =============================================================================
    # 5. COURSE ENROLLMENT & COMPLETION ANALYTICS
    # =============================================================================
    def get_course_analytics():
        # Most popular courses
        popular_courses = Enrollment.objects.values(
            'allocationid__coursecode__coursename',
            'allocationid__coursecode__coursecode',
            'allocationid__coursecode__credithours'
        ).annotate(
            enrollment_count=Count('enrollmentid'),
            completion_count=Count('enrollmentid', filter=Q(status='Completed')),
            active_count=Count('enrollmentid', filter=Q(status='Active')),
            dropout_count=Count('enrollmentid', filter=Q(status='Dropped')),
            avg_gpa=Avg('result__coursegpa')
        ).order_by('-enrollment_count')[:15]

        # Calculate completion rates
        for course in popular_courses:
            if course['enrollment_count'] > 0:
                course['completion_rate'] = (course['completion_count'] / course['enrollment_count']) * 100
                course['dropout_rate'] = (course['dropout_count'] / course['enrollment_count']) * 100
            else:
                course['completion_rate'] = course['dropout_rate'] = 0

        # Challenging courses (low average GPA)
        challenging_courses = Enrollment.objects.values(
            'allocationid__coursecode__coursename',
            'allocationid__coursecode__coursecode'
        ).annotate(
            avg_gpa=Avg('result__coursegpa'),
            enrollment_count=Count('enrollmentid')
        ).filter(
            enrollment_count__gte=5  # At least 5 enrollments
        ).order_by('avg_gpa')[:10]

        return {
            'popular_courses': popular_courses,
            'challenging_courses': challenging_courses
        }

    # =============================================================================
    # 6. ENROLLMENT TRENDS & PATTERNS
    # =============================================================================
    def get_enrollment_trends():
        # Monthly enrollment trends (last 12 months)
        twelve_months_ago = timezone.now() - timedelta(days=365)
        monthly_enrollments = []

        for i in range(12):
            month_start = twelve_months_ago + timedelta(days=30 * i)
            month_end = month_start + timedelta(days=30)

            month_data = {
                'month': month_start.strftime('%Y-%m'),
                'month_name': month_start.strftime('%B %Y'),
                'enrollments': Enrollment.objects.filter(
                    enrollmentdate__gte=month_start,
                    enrollmentdate__lt=month_end
                ).count(),
                'new_students': Student.objects.filter(
                    enrollment__enrollmentdate__gte=month_start,
                    enrollment__enrollmentdate__lt=month_end
                ).distinct().count()
            }
            monthly_enrollments.append(month_data)

        # Semester-wise enrollment patterns
        semester_patterns = Enrollment.objects.values(
            'allocationid__coursecode__semesterdetails__semesterid__semesterno',
            'allocationid__coursecode__semesterdetails__semesterid__session'
        ).annotate(
            enrollment_count=Count('enrollmentid'),
            unique_students=Count('studentid', distinct=True),
            completion_rate=Count('enrollmentid', filter=Q(status='Completed')) * 100.0 / Count('enrollmentid')
        ).order_by('allocationid__coursecode__semesterdetails__semesterid__semesterno')

        return {
            'monthly_enrollments': monthly_enrollments,
            'semester_patterns': semester_patterns
        }

    # =============================================================================
    # 7. AT-RISK STUDENT IDENTIFICATION
    # =============================================================================
    def identify_at_risk_students():
        # Students with low GPA (< 2.0)
        low_gpa_students = Student.objects.annotate(
            avg_gpa=Avg('enrollment__result__coursegpa'),
            total_courses=Count('enrollment__result')
        ).filter(
            avg_gpa__lt=2.0,
            status='Enrolled',
            total_courses__gte=2  # At least 2 courses completed
        )

        # Students with multiple course failures
        failing_students = Student.objects.annotate(
            failed_courses=Count('enrollment', filter=Q(enrollment__result__coursegpa__lt=2.0)),
            total_courses=Count('enrollment__result')
        ).filter(
            failed_courses__gte=2,
            status='Enrolled'
        )

        # Students with no recent enrollments (inactive)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        inactive_students = Student.objects.filter(
            status='Enrolled'
        ).exclude(
            enrollment__enrollmentdate__gte=thirty_days_ago
        ).annotate(
            last_enrollment=Max('enrollment__enrollmentdate')
        )

        # Students with high dropout rates in courses
        high_dropout_students = Student.objects.annotate(
            dropped_courses=Count('enrollment', filter=Q(enrollment__status='Dropped')),
            total_enrollments=Count('enrollment')
        ).filter(
            total_enrollments__gte=3,
            dropped_courses__gte=2,
            status='Enrolled'
        )

        return {
            'low_gpa_students': low_gpa_students,
            'failing_students': failing_students,
            'inactive_students': inactive_students,
            'high_dropout_students': high_dropout_students
        }

    # =============================================================================
    # 8. SEMESTER PROGRESSION TRACKING
    # =============================================================================
    def track_semester_progression():
        # Track student progress through semesters
        semester_completion = Transcript.objects.values(
            'semesterid__semesterno',
            'semesterid__session',
            'studentid__classid__batchyear'
        ).annotate(
            students_completed=Count('studentid'),
            avg_semester_gpa=Avg('semestergpa'),
            avg_credits=Avg('totalcredits')
        ).order_by('semesterid__semesterno')

        return semester_completion

    # =============================================================================
    # 9. COMPARATIVE ANALYTICS
    # =============================================================================
    def generate_comparative_analytics():
        # Compare performance across batches
        batch_comparison = Student.objects.values(
            'classid__batchyear'
        ).annotate(
            student_count=Count('studentid'),
            avg_gpa=Avg('enrollment__result__coursegpa'),
            graduation_rate=Count('studentid', filter=Q(status='Graduated')) * 100.0 / Count('studentid')
        ).order_by('-classid__batchyear')

        # Compare departments
        dept_comparison = get_department_analytics()

        return {
            'batch_comparison': batch_comparison,
            'department_comparison': dept_comparison
        }

    # =============================================================================
    # 10. GENERATE ACTIONABLE INSIGHTS
    # =============================================================================
    def generate_insights(analytics_data):
        insights = {
            'alerts': [],
            'recommendations': [],
            'key_findings': [],
            'action_items': []
        }

        # Calculate key metrics
        total = analytics_data['total_students']
        dropout_rate = (analytics_data['dropped_students'] / total * 100) if total > 0 else 0

        # Generate alerts
        if dropout_rate > 15:
            insights['alerts'].append(f"High dropout rate: {dropout_rate:.1f}% of students have dropped")

        if analytics_data['at_risk']['low_gpa_students'].count() > 0:
            insights['alerts'].append(
                f"{analytics_data['at_risk']['low_gpa_students'].count()} students have GPA below 2.0")

        # Generate recommendations
        if analytics_data['at_risk']['inactive_students'].count() > 5:
            insights['recommendations'].append("Implement early intervention system for inactive students")

        if dropout_rate > 10:
            insights['recommendations'].append("Review and improve student support services")

        # Key findings
        if analytics_data['performance']['gpa_stats']['overall_avg_gpa']:
            avg_gpa = analytics_data['performance']['gpa_stats']['overall_avg_gpa']
            insights['key_findings'].append(f"Overall student GPA: {avg_gpa:.2f}")

        return insights

    # =============================================================================
    # COLLECT ALL DATA
    # =============================================================================
    batch_analytics = get_batch_analytics()
    program_analytics = get_program_analytics()
    department_analytics = get_department_analytics()
    performance_analytics = get_performance_analytics()
    course_analytics = get_course_analytics()
    enrollment_trends = get_enrollment_trends()
    at_risk_analysis = identify_at_risk_students()
    semester_progression = track_semester_progression()
    comparative_analytics = generate_comparative_analytics()

    # Prepare analytics data for insights
    analytics_data = {
        'total_students': total_students,
        'dropped_students': dropped_students,
        'at_risk': at_risk_analysis,
        'performance': performance_analytics
    }

    insights = generate_insights(analytics_data)

    # =============================================================================
    # CONTEXT FOR TEMPLATE
    # =============================================================================
    context = {
        # Basic statistics
        'total_students': total_students,
        'active_students': active_students,
        'graduated_students': graduated_students,
        'dropped_students': dropped_students,

        # Key rates
        'dropout_rate': (dropped_students / total_students * 100) if total_students > 0 else 0,
        'graduation_rate': (graduated_students / total_students * 100) if total_students > 0 else 0,
        'retention_rate': (active_students / total_students * 100) if total_students > 0 else 0,

        # Detailed analytics
        'batch_analytics': batch_analytics,
        'program_analytics': program_analytics,
        'department_analytics': department_analytics,
        'performance_analytics': performance_analytics,
        'course_analytics': course_analytics,
        'enrollment_trends': enrollment_trends,
        'at_risk_analysis': at_risk_analysis,
        'semester_progression': semester_progression,
        'comparative_analytics': comparative_analytics,

        # Insights and recommendations
        'insights': insights,

        # Additional metrics for dashboard
        'avg_students_per_batch': total_students / batch_analytics.count() if batch_analytics.count() > 0 else 0,
        'total_programs': program_analytics.count(),
        'total_departments': department_analytics.count(),
    }

    return render(request, 'admin/student_analytics_report.html', context)


@login_required
@user_passes_test(is_admin)
def semester_performance_report(request):
    """Comprehensive analytics report for all semesters grouped by class"""

    def get_all_classes_with_semesters():
        """Get all classes that have semester details"""
        classes_with_semesters = Class.objects.filter(
            semesterdetails__isnull=False
        ).select_related('programid__departmentid').distinct().order_by(
            'programid__departmentid__departmentname',
            'programid__programname',
            '-batchyear'
        )
        return classes_with_semesters

    def get_semester_performance_for_class(class_obj):
        """Get performance data for all semesters of a specific class"""
        # Get all semesters linked to this class
        semester_details = Semesterdetails.objects.filter(
            classid=class_obj
        ).select_related('semesterid').values_list('semesterid', flat=True).distinct()

        semesters = Semester.objects.filter(
            semesterid__in=semester_details
        ).order_by('semesterno')

        semester_performance = []

        for semester in semesters:
            # Get courses for this semester and class
            courses = Semesterdetails.objects.filter(
                semesterid=semester,
                classid=class_obj
            ).select_related('coursecode')

            course_codes = [sd.coursecode for sd in courses]

            # Get course allocations
            allocations = Courseallocation.objects.filter(
                coursecode__in=course_codes
            )

            # Get enrollments for this class only
            enrollments = Enrollment.objects.filter(
                allocationid__in=allocations,
                studentid__classid=class_obj
            )

            # Get results
            results = Result.objects.filter(
                enrollmentid__in=enrollments
            )

            # Calculate performance metrics
            performance_data = {
                'semester': semester,
                'total_courses': courses.count(),
                'total_enrollments': enrollments.count(),
                'total_results': results.count(),
                'completion_rate': 0,
                'average_marks': 0,
                'average_gpa': 0,
                'pass_rate': 0,
                'excellent_rate': 0,
                'allocations_count': allocations.count(),
            }

            if results.exists():
                # Calculate metrics using obtained marks and GPA
                marks_data = [r.obtainedmarks for r in results if r.obtainedmarks is not None]
                gpa_data = [r.coursegpa for r in results if r.coursegpa is not None]

                if marks_data:
                    import statistics
                    performance_data.update({
                        'completion_rate': round((results.count() / enrollments.count()) * 100,
                                                 1) if enrollments.count() > 0 else 0,
                        'average_marks': round(statistics.mean(marks_data), 2),
                        'pass_rate': round((len([m for m in marks_data if m >= 50]) / len(marks_data)) * 100, 1),
                        'excellent_rate': round((len([m for m in marks_data if m >= 80]) / len(marks_data)) * 100, 1),
                    })

                if gpa_data:
                    performance_data['average_gpa'] = round(statistics.mean(gpa_data), 2)

            # Get transcript data for completed semesters
            transcript_data = None
            if semester.status == 'Completed':
                transcripts = Transcript.objects.filter(
                    semesterid=semester,
                    studentid__classid=class_obj
                )

                if transcripts.exists():
                    transcript_gpas = [t.semestergpa for t in transcripts if t.semestergpa is not None]
                    if transcript_gpas:
                        import statistics
                        transcript_data = {
                            'count': transcripts.count(),
                            'average_gpa': round(statistics.mean(transcript_gpas), 2),
                            'highest_gpa': round(max(transcript_gpas), 2),
                            'lowest_gpa': round(min(transcript_gpas), 2),
                        }

            performance_data['transcript_data'] = transcript_data
            semester_performance.append(performance_data)

        return semester_performance

    def calculate_class_overall_stats(semester_performance):
        """Calculate overall statistics for a class across all semesters"""
        if not semester_performance:
            return None

        total_enrollments = sum(sp['total_enrollments'] for sp in semester_performance)
        total_results = sum(sp['total_results'] for sp in semester_performance)
        completed_semesters = len([sp for sp in semester_performance if sp['semester'].status == 'Completed'])
        active_semesters = len([sp for sp in semester_performance if sp['semester'].status == 'Active'])

        # Calculate weighted averages
        weighted_marks = []
        weighted_gpas = []
        total_courses = sum(sp['total_courses'] for sp in semester_performance)

        for sp in semester_performance:
            if sp['total_results'] > 0:
                # Weight by number of results
                weight = sp['total_results']
                if sp['average_marks'] > 0:
                    weighted_marks.extend([sp['average_marks']] * weight)
                if sp['average_gpa'] > 0:
                    weighted_gpas.extend([sp['average_gpa']] * weight)

        overall_stats = {
            'total_semesters': len(semester_performance),
            'completed_semesters': completed_semesters,
            'active_semesters': active_semesters,
            'total_courses': total_courses,
            'total_enrollments': total_enrollments,
            'total_results': total_results,
            'overall_completion_rate': round((total_results / total_enrollments) * 100,
                                             1) if total_enrollments > 0 else 0,
        }

        if weighted_marks:
            import statistics
            overall_stats['overall_average_marks'] = round(statistics.mean(weighted_marks), 2)

        if weighted_gpas:
            import statistics
            overall_stats['overall_average_gpa'] = round(statistics.mean(weighted_gpas), 2)

        return overall_stats

    def prepare_chart_data_for_class(semester_performance):
        """Prepare chart data for class performance visualization"""
        chart_data = {
            'semester_labels': [],
            'completion_rates': [],
            'average_marks': [],
            'average_gpas': [],
            'enrollment_counts': [],
            'transcript_gpas': []
        }

        for sp in semester_performance:
            chart_data['semester_labels'].append(f"Sem {sp['semester'].semesterno}")
            chart_data['completion_rates'].append(sp['completion_rate'])
            chart_data['average_marks'].append(sp['average_marks'])
            chart_data['average_gpas'].append(sp['average_gpa'])
            chart_data['enrollment_counts'].append(sp['total_enrollments'])

            # Add transcript GPA if available
            if sp['transcript_data']:
                chart_data['transcript_gpas'].append(sp['transcript_data']['average_gpa'])
            else:
                chart_data['transcript_gpas'].append(0)

        return chart_data

    # Main execution
    classes = get_all_classes_with_semesters()

    class_reports = []
    overall_system_stats = {
        'total_classes': 0,
        'total_semesters': 0,
        'total_completed_semesters': 0,
        'total_students': 0,
        'total_enrollments': 0,
        'departments': set(),
        'programs': set()
    }

    for class_obj in classes:
        # Get semester performance for this class
        semester_performance = get_semester_performance_for_class(class_obj)

        if semester_performance:  # Only include classes with semester data
            # Calculate overall stats for this class
            overall_stats = calculate_class_overall_stats(semester_performance)

            # Prepare chart data
            chart_data = prepare_chart_data_for_class(semester_performance)

            # Get student count for this class
            from StudentModule.models import Student
            student_count = Student.objects.filter(classid=class_obj).count()

            # Current semester info (using existing logic from class_list view)
            current_semester = None
            semester_status = "No Active Semester"

            active_enrollments = Enrollment.objects.filter(
                studentid__classid=class_obj,
                status='Active',
                allocationid__status='Ongoing'
            ).select_related('allocationid__coursecode')

            if active_enrollments.exists():
                semester_counts = {}
                for enrollment in active_enrollments:
                    course = enrollment.allocationid.coursecode
                    semester_detail = Semesterdetails.objects.filter(
                        coursecode=course,
                        classid=class_obj
                    ).select_related('semesterid').first()

                    if semester_detail and semester_detail.semesterid.status == 'Active':
                        sem_no = semester_detail.semesterid.semesterno
                        if sem_no not in semester_counts:
                            semester_counts[sem_no] = {
                                'count': 0,
                                'semester': semester_detail.semesterid
                            }
                        semester_counts[sem_no]['count'] += 1

                if semester_counts:
                    max_sem_no = max(semester_counts.keys(), key=lambda x: semester_counts[x]['count'])
                    current_semester = semester_counts[max_sem_no]['semester']
                    semester_status = f"Semester {max_sem_no}"

            class_report = {
                'class_obj': class_obj,
                'class_display': f"{class_obj.programid.programid}-{class_obj.batchyear}",
                'department': class_obj.programid.departmentid,
                'program': class_obj.programid,
                'student_count': student_count,
                'current_semester': current_semester,
                'semester_status': semester_status,
                'semester_performance': semester_performance,
                'overall_stats': overall_stats,
                'chart_data': chart_data
            }

            class_reports.append(class_report)

            # Update system-wide stats
            overall_system_stats['total_classes'] += 1
            overall_system_stats['total_semesters'] += overall_stats['total_semesters']
            overall_system_stats['total_completed_semesters'] += overall_stats['completed_semesters']
            overall_system_stats['total_students'] += student_count
            overall_system_stats['total_enrollments'] += overall_stats['total_enrollments']
            overall_system_stats['departments'].add(class_obj.programid.departmentid.departmentname)
            overall_system_stats['programs'].add(class_obj.programid.programid)

    # Convert sets to counts
    overall_system_stats['total_departments'] = len(overall_system_stats['departments'])
    overall_system_stats['total_programs'] = len(overall_system_stats['programs'])
    overall_system_stats.pop('departments')
    overall_system_stats.pop('programs')

    # Group by department for better organization
    department_groups = {}
    for class_report in class_reports:
        dept_name = class_report['department'].departmentname
        if dept_name not in department_groups:
            department_groups[dept_name] = {
                'department': class_report['department'],
                'classes': []
            }
        department_groups[dept_name]['classes'].append(class_report)

    # Sort classes within each department
    for dept_data in department_groups.values():
        dept_data['classes'].sort(key=lambda x: (x['program'].programname, -int(x['class_obj'].batchyear)))

    context = {
        'department_groups': department_groups,
        'overall_system_stats': overall_system_stats,
        'total_departments': len(department_groups),
    }

    return render(request, 'admin/semester_performance_report.html', context)


# ===========================================
# AUDIT TRAIL MANAGEMENT (Form-based) - FIXED field names
# ===========================================

# Add these views to your views.py file

from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
import csv
import json


@login_required
@user_passes_test(is_admin)
def audit_trail_detail(request, audit_id):
    """Get detailed information about a specific audit record"""
    try:
        audit_record = Audittrail.objects.select_related('userid').get(auditid=audit_id)

        data = {
            'id': audit_record.auditid,
            'timestamp': audit_record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'action': audit_record.actiontype,
            'table': audit_record.entityname,
            'ipaddress': audit_record.ipaddress,
            'useragent' : audit_record.useragent,
            'user': {
                'name': f"{audit_record.userid.fname} {audit_record.userid.lname}" if audit_record.userid else "System",
                'email': audit_record.userid.institutionalemail if audit_record.userid else None
            },

        }
        print(data)

        return JsonResponse(data)

    except Audittrail.DoesNotExist:
        return JsonResponse({'error': 'Audit record not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



# Update your existing audit_trail_list view to handle export
@login_required
@user_passes_test(is_admin)
def audit_trail_list(request):
    """View audit trail with filtering and export functionality"""

    # Check if this is an export request
    if request.GET.get('export') == 'csv':
        return export_audit_trail(request)

    audit_records = Audittrail.objects.select_related('userid').all()
    today = timezone.now().date()
    today_count = Audittrail.objects.filter(
        timestamp__date=today
    ).count()

    # Filter by action
    action = request.GET.get('action')
    if action and action != 'None' and action.strip():
        audit_records = audit_records.filter(actiontype=action)

    # Filter by table
    table = request.GET.get('table')
    if table and table != 'None' and table.strip():
        audit_records = audit_records.filter(entityname=table)

    # Filter by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and start_date != 'None' and start_date.strip():
        try:
            audit_records = audit_records.filter(timestamp__gte=start_date)
        except ValidationError:
            pass
    if end_date and end_date != 'None' and end_date.strip():
        try:
            audit_records = audit_records.filter(timestamp__lte=end_date)
        except ValidationError:
            pass

    # Filter by user
    user_filter = request.GET.get('user')
    if user_filter and user_filter != 'None' and user_filter.strip():
        audit_records = audit_records.filter(userid__fname__icontains=user_filter)

    # Pagination
    paginator = Paginator(audit_records.order_by('-timestamp'), 50)
    page = request.GET.get('page')
    audit_records = paginator.get_page(page)

    # Get distinct values for filters
    distinct_actions = Audittrail.objects.values_list('actiontype', flat=True).distinct()
    distinct_tables = Audittrail.objects.values_list('entityname', flat=True).distinct()

    context = {
        'audit_records': audit_records,
        'today_count' : today_count,
        'distinct_actions': distinct_actions,
        'distinct_tables': distinct_tables,
        'current_filters': {
            'action': action,
            'table': table,
            'start_date': start_date,
            'end_date': end_date,
            'user': user_filter,
        }
    }

    return render(request, 'admin/audit_trail_list.html', context)

@login_required
@user_passes_test(is_admin)
def export_audit_trail(request):
    """Export audit trail data as CSV"""
    # Get the same filtered queryset as the list view
    audit_records = Audittrail.objects.select_related('userid').all()

    # Apply the same filters as in audit_trail_list
    action = request.GET.get('action')
    if action:
        audit_records = audit_records.filter(actiontype=action)

    table = request.GET.get('table')
    if table:
        audit_records = audit_records.filter(entityname=table)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        audit_records = audit_records.filter(timestamp__gte=start_date)
    if end_date:
        audit_records = audit_records.filter(timestamp__lte=end_date)

    user_filter = request.GET.get('user')
    if user_filter:
        audit_records = audit_records.filter(userid__fname__icontains=user_filter)

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="audit_trail.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Timestamp', 'User', 'User Email', 'Action', 'Table',
        'Record ID', 'Old Values', 'New Values'
    ])

    for record in audit_records.order_by('-timestamp'):
        user_name = f"{record.userid.fname} {record.userid.lname}" if record.userid else "System"
        user_email = record.userid.email if record.userid else ""

        writer.writerow([
            record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            user_name,
            user_email,
            record.actiontype,
            record.entityname,
            record.entityid or '',
            record.oldvalues or '',
            record.newvalues or '',
        ])

    return response

# ===========================================
# BULK OPERATIONS (Form-based)
# ===========================================

@login_required
@user_passes_test(is_admin)
def bulk_operations(request):
    """Main bulk operations page"""
    total_students = Student.objects.count()
    total_faculty = Faculty.objects.count()
    active_allocations = Courseallocation.objects.filter(
        status='Ongoing',
    ).count()
    active_enrollments = Enrollment.objects.filter(
        status='Active',
    ).count()
    return render(request, 'admin/bulk_operations.html',{
        'total_students': total_students,
        'total_faculty': total_faculty,
        'active_allocations': active_allocations,
        'active_enrollments': active_enrollments,
    })


@login_required
@user_passes_test(is_admin)
def bulk_student_operations(request):
    """Perform bulk operations on students"""
    if request.method == 'POST':
        operation = request.POST.get('operation')
        student_ids = request.POST.getlist('student_ids')

        try:
            if operation == 'activate':
                Student.objects.filter(studentid__personid__in=student_ids).update(status='Enrolled')
                messages.success(request, f'Activated {len(student_ids)} students')

            elif operation == 'deactivate':
                Student.objects.filter(studentid__personid__in=student_ids).update(status='Dropped')
                messages.success(request, f'Deactivated {len(student_ids)} students')

            elif operation == 'graduate':
                Student.objects.filter(studentid__personid__in=student_ids).update(status='Graduated')
                messages.success(request, f'Marked {len(student_ids)} students as graduated')

            else:
                messages.error(request, 'Invalid operation')

        except Exception as e:
            messages.error(request, f'Bulk operation failed: {str(e)}')

        return redirect('management:bulk_student_operations')

    # Get students with filters for bulk operations
    students = Student.objects.select_related('studentid', 'programid').all()

    # Apply filters
    program_filter = request.GET.get('program')
    if program_filter:
        students = students.filter(programid=program_filter)

    status_filter = request.GET.get('status')
    if status_filter:
        students = students.filter(status=status_filter)

    # Pagination
    paginator = Paginator(students, 50)
    page = request.GET.get('page')
    students = paginator.get_page(page)

    context = {
        'students': students,
        'programs': Program.objects.all(),
        'current_filters': {
            'program': program_filter,
            'status': status_filter,
        }
    }

    return render(request, 'admin/bulk_student_operations.html', context)


# ===========================================
# SYSTEM UTILITIES (Form-based)
# ===========================================

@login_required
@user_passes_test(is_admin)
def system_health_check(request):
    """Check system health and provide diagnostics"""
    health_status = {
        'database': 'OK',
        'models': 'OK',
        'issues': []
    }

    try:
        # Test database connectivity
        Student.objects.count()
        Faculty.objects.count()
        Department.objects.count()
    except Exception as e:
        health_status['database'] = 'ERROR'
        health_status['issues'].append(f'Database connectivity issue: {str(e)}')

    # Check for orphaned records
    orphaned_issues = []

    try:
        # Students without valid programs
        orphaned_students = Student.objects.filter(programid__isnull=True).count()
        if orphaned_students > 0:
            orphaned_issues.append(f'{orphaned_students} students without valid programs')

        # Faculty without departments
        orphaned_faculty = Faculty.objects.filter(departmentid__isnull=True).count()
        if orphaned_faculty > 0:
            orphaned_issues.append(f'{orphaned_faculty} faculty without departments')

        # Enrollments with invalid allocations
        orphaned_enrollments = Enrollment.objects.filter(allocationid__isnull=True).count()
        if orphaned_enrollments > 0:
            orphaned_issues.append(f'{orphaned_enrollments} enrollments with invalid allocations')

    except Exception as e:
        orphaned_issues.append(f'Error checking orphaned records: {str(e)}')

    if orphaned_issues:
        health_status['models'] = 'WARNING'
        health_status['issues'].extend(orphaned_issues)

    # Check for recent activity (using actual model fields)
    recent_activity = {
        'enrollments_last_week': Enrollment.objects.filter(
            enrollmentdate__gte=timezone.now() - timedelta(days=7)
        ).count(),
        'new_faculty_last_month': Faculty.objects.filter(
            joiningdate__gte=timezone.now() - timedelta(days=30)
        ).count(),
        'recent_lectures': Lecture.objects.filter(
            startingtime__gte=timezone.now() - timedelta(days=7)
        ).count(),
    }

    context = {
        'health_status': health_status,
        'recent_activity': recent_activity,
        'timestamp': timezone.now(),
    }

    return render(request, 'admin/system_health.html', context)



# ===========================================
# ANALYTICS AND REPORTS (Form-based)
# ===========================================

@login_required
@user_passes_test(is_admin)
def analytics_dashboard(request):
    """Main analytics dashboard - FIXED field names"""
    # Student analytics
    total_students = Student.objects.count()
    students_by_status = Student.objects.values('status').annotate(count=Count('studentid')).order_by('status')
    students_by_program = Student.objects.values('programid__programname').annotate(count=Count('studentid')).order_by(
        '-count')[:10]

    # Faculty analytics
    total_faculty = Faculty.objects.count()
    faculty_by_department = Faculty.objects.values('departmentid__departmentname').annotate(
        count=Count('employeeid')).order_by('-count')

    # Performance analytics - FIXED field names
    try:
        avg_gpa = Result.objects.aggregate(avg_gpa=Avg('coursegpa'))['avg_gpa'] or 0.0  # FIXED: coursegpa
        # Removed grade_distribution since grade field doesn't exist
        high_performers = Result.objects.filter(coursegpa__gte=3.5).count()  # FIXED: coursegpa
    except:
        avg_gpa = 0.0
        high_performers = 0

    # Enrollment trends (last 6 months)
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_enrollments = []

    for i in range(6):
        month_start = six_months_ago + timedelta(days=30 * i)
        month_end = month_start + timedelta(days=30)

        enrollments_count = Enrollment.objects.filter(
            enrollmentdate__gte=month_start,
            enrollmentdate__lt=month_end
        ).count()

        monthly_enrollments.append({
            'month': month_start.strftime('%Y-%m'),
            'enrollments': enrollments_count
        })

    context = {
        'student_analytics': {
            'total': total_students,
            'by_status': students_by_status,
            'by_program': students_by_program,
        },
        'faculty_analytics': {
            'total': total_faculty,
            'by_department': faculty_by_department,
        },
        'performance_analytics': {
            'avg_gpa': round(float(avg_gpa), 2) if avg_gpa else 0.0,
            'high_performers': high_performers,
        },
        'enrollment_trends': monthly_enrollments,
    }

    return render(request, context)




# ===========================================
# UTILITIES AND HELPER FUNCTIONS
# ===========================================

@login_required
@user_passes_test(is_admin)
def data_integrity_check(request):
    """Check data integrity across the system"""
    issues = []

    # Check for orphaned records
    try:
        # Students without programs
        orphaned_students = Student.objects.filter(programid__isnull=True)
        if orphaned_students.exists():
            issues.append({
                'type': 'orphaned_students',
                'count': orphaned_students.count(),
                'description': 'Students without assigned programs',
                'records': orphaned_students[:10]  # Show first 10
            })

        # Faculty without departments
        orphaned_faculty = Faculty.objects.filter(departmentid__isnull=True)
        if orphaned_faculty.exists():
            issues.append({
                'type': 'orphaned_faculty',
                'count': orphaned_faculty.count(),
                'description': 'Faculty without assigned departments',
                'records': orphaned_faculty[:10]
            })

        # Enrollments without allocations
        orphaned_enrollments = Enrollment.objects.filter(allocationid__isnull=True)
        if orphaned_enrollments.exists():
            issues.append({
                'type': 'orphaned_enrollments',
                'count': orphaned_enrollments.count(),
                'description': 'Enrollments without valid course allocations',
                'records': orphaned_enrollments[:10]
            })

        # Check for duplicate person IDs
        duplicate_persons = Person.objects.values('personid').annotate(count=Count('personid')).filter(count__gt=1)
        if duplicate_persons.exists():
            issues.append({
                'type': 'duplicate_persons',
                'count': duplicate_persons.count(),
                'description': 'Duplicate person IDs found',
                'records': duplicate_persons[:10]
            })

    except Exception as e:
        issues.append({
            'type': 'error',
            'description': f'Error during integrity check: {str(e)}'
        })

    context = {
        'issues': issues,
        'total_issues': len(issues),
        'check_timestamp': timezone.now()
    }

    return render(request, 'admin/data_integrity_check.html', context)


@login_required
@user_passes_test(is_admin)
def fix_data_issue(request, issue_type):
    """Fix specific data integrity issues"""
    if request.method == 'POST':
        try:
            if issue_type == 'orphaned_students':
                # Logic to fix orphaned students
                # This would depend on your business rules
                messages.success(request, 'Orphaned students issue resolution initiated')

            elif issue_type == 'orphaned_faculty':
                # Logic to fix orphaned faculty
                messages.success(request, 'Orphaned faculty issue resolution initiated')

            elif issue_type == 'orphaned_enrollments':
                # Logic to fix orphaned enrollments
                messages.success(request, 'Orphaned enrollments issue resolution initiated')

            else:
                messages.error(request, 'Unknown issue type')

        except Exception as e:
            messages.error(request, f'Error fixing issue: {str(e)}')

    return redirect('management:data_integrity_check')





# ===========================================
# MAIN IMPORT PAGE
# ===========================================

@login_required
@user_passes_test(is_admin)
def data_import_page(request):
    """Main data import page - reuses your bulk_operations pattern"""
    # Reuse your existing stats calculation pattern
    total_students = Student.objects.count()
    total_faculty = Faculty.objects.count()
    active_allocations = Courseallocation.objects.filter(status='Active').count()
    active_enrollments = Enrollment.objects.filter(status='Active').count()

    # Recent imports from audit trail (using your existing model)
    recent_imports = Audittrail.objects.filter(
        actiontype='CSV_IMPORT'  # We'll use a special action type for imports
    ).select_related('userid').order_by('-timestamp')[:10]

    context = {
        'total_students': total_students,
        'total_faculty': total_faculty,
        'active_allocations': active_allocations,
        'active_enrollments': active_enrollments,
        'recent_imports': recent_imports,
    }

    return render(request, 'admin/data_import.html', context)


# ===========================================
# CSV PROCESSING ENGINE
# ===========================================

@login_required
@user_passes_test(is_admin)
def process_csv_import(request):
    """Handle CSV file upload and processing - reuses your JSON response pattern"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

    try:
        # Get uploaded file and import type
        csv_file = request.FILES.get('csv_file')
        import_type = request.POST.get('import_type')

        if not csv_file or not import_type:
            return JsonResponse({
                'success': False,
                'error': 'Missing file or import type'
            }, status=400)

        # Validate file
        if not csv_file.name.endswith('.csv'):
            return JsonResponse({
                'success': False,
                'error': 'File must be a CSV'
            }, status=400)

        if csv_file.size > 10 * 1024 * 1024:  # 10MB limit
            return JsonResponse({
                'success': False,
                'error': 'File size must be less than 10MB'
            }, status=400)

        # Get current admin (reuse your admin profile pattern)
        admin = get_object_or_404(Admin, employeeid__institutionalemail=request.user.username)
        admin_person = admin.employeeid

        # Track import start in audit trail (using your existing model)
        audit_start = Audittrail.objects.create(
            userid=admin_person,
            actiontype='CSV_IMPORT',
            entityname=f'{import_type}_import',
            description=f'Started CSV import of {csv_file.name}',
            timestamp=timezone.now()
        )

        # Process CSV based on type
        if import_type == 'students':
            result = process_student_csv(csv_file, admin_person)
        elif import_type == 'faculty':
            result = process_faculty_csv(csv_file, admin_person)
        else:
            return JsonResponse({
                'success': False,
                'error': 'Invalid import type'
            }, status=400)

        # Track import completion in audit trail (using your existing model)
        status = 'completed' if result['successful'] > 0 else 'failed'
        description = f'CSV import completed: {result["successful"]}/{result["total_processed"]} successful'

        Audittrail.objects.create(
            userid=admin_person,
            actiontype='CSV_IMPORT',
            entityname=f'{import_type}_import',
            description=description,
            timestamp=timezone.now()
        )

        # Return result (reuse your JSON response pattern)
        return JsonResponse({
            'success': True,
            'total_processed': result['total_processed'],
            'successful': result['successful'],
            'failed': result['failed'],
            'skipped': result['skipped'],
            'errors': result.get('errors', [])[:10]  # Limit errors shown
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Import failed: {str(e)}'
        }, status=500)


# ===========================================
# STUDENT CSV PROCESSING
# ===========================================

def process_student_csv(csv_file, admin_person):
    """Process student CSV - reuses your StudentForm validation patterns"""
    results = {
        'total_processed': 0,
        'successful': 0,
        'failed': 0,
        'skipped': 0,
        'errors': []
    }

    try:
        # Read CSV file
        csv_data = csv_file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_data))

        # Expected columns for students (matching your StudentForm fields)
        required_columns = [
            'studentid', 'fname', 'lname', 'institutionalemail',
            'cnumber', 'programid', 'classid'
        ]

        # Optional columns that match your form
        optional_columns = [
            'personalemail', 'cnic', 'gender', 'dob', 'status',
            'country', 'province', 'city', 'zipcode', 'streetaddress'
        ]

        # Validate CSV headers
        if not all(col in csv_reader.fieldnames for col in required_columns):
            missing = [col for col in required_columns if col not in csv_reader.fieldnames]
            results['errors'].append({
                'row': 0,
                'message': f'Missing required columns: {", ".join(missing)}'
            })
            return results

        # Process each row
        for row_num, row in enumerate(csv_reader, start=1):
            results['total_processed'] += 1

            try:
                # Check if student already exists (reuse your validation pattern)
                if Person.objects.filter(personid=row['studentid']).exists():
                    results['skipped'] += 1
                    results['errors'].append({
                        'row': row_num,
                        'message': f'Student ID {row["studentid"]} already exists'
                    })
                    continue

                # Prepare data for StudentForm (matching your exact field names)
                form_data = {
                    'studentid': row['studentid'],
                    'fname': row['fname'],
                    'lname': row['lname'],
                    'institutionalemail': row['institutionalemail'],
                    'personalemail': row.get('personalemail', ''),
                    'cnic': row.get('cnic', ''),
                    'gender': row.get('gender', 'M'),
                    'dob': parse_date(row.get('dob', '')),
                    'cnumber': row.get('cnumber', ''),
                    'programid': get_program_by_id(row.get('programid', '')),
                    'classid': get_class_by_id(row.get('classid', '')),
                    'status': row.get('status', 'Enrolled'),
                    # Address fields
                    'country': row.get('country', 'Pakistan'),
                    'province': row.get('province', ''),
                    'city': row.get('city', ''),
                    'zipcode': parse_int(row.get('zipcode', '0')),
                    'streetaddress': row.get('streetaddress', ''),
                }

                # Create StudentForm instance (reuse your form validation)
                form = StudentForm(data=form_data)

                if form.is_valid():
                    # Convert CSV qualifications to POST-like format for your existing function
                    post_like_data = convert_csv_qualifications_to_post_format(row)

                    # Use your existing qualification extraction function
                    qualifications_data = student_views.extract_qualifications_from_post(post_like_data)

                    # Save student with qualifications (reuse your save method)
                    student = form.save_with_qualifications(qualifications_data, commit=True)
                    results['successful'] += 1
                else:
                    # Collect form errors
                    error_messages = []
                    for field, errors in form.errors.items():
                        error_messages.extend([f"{field}: {error}" for error in errors])

                    results['failed'] += 1
                    results['errors'].append({
                        'row': row_num,
                        'message': '; '.join(error_messages)
                    })

            except Exception as e:
                results['failed'] += 1
                results['errors'].append({
                    'row': row_num,
                    'message': f'Processing error: {str(e)}'
                })

        return results

    except Exception as e:
        results['errors'].append({
            'row': 0,
            'message': f'CSV parsing error: {str(e)}'
        })
        return results


# ===========================================
# FACULTY CSV PROCESSING (Updated)
# ===========================================

def process_faculty_csv(csv_file, admin_person):
    """Process faculty CSV - reuses your FacultyForm validation patterns"""
    results = {
        'total_processed': 0,
        'successful': 0,
        'failed': 0,
        'skipped': 0,
        'errors': []
    }

    try:
        # Read CSV file
        csv_data = csv_file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_data))

        # Expected columns for faculty (matching your FacultyForm fields)
        required_columns = [
            'employeeid', 'fname', 'lname', 'institutionalemail',
            'cnumber', 'departmentid', 'designation', 'joiningdate'
        ]

        # Optional columns that match your form
        optional_columns = [
            'personalemail', 'cnic', 'gender', 'dob',
            'country', 'province', 'city', 'zipcode', 'streetaddress'
        ]

        # Validate CSV headers
        if not all(col in csv_reader.fieldnames for col in required_columns):
            missing = [col for col in required_columns if col not in csv_reader.fieldnames]
            results['errors'].append({
                'row': 0,
                'message': f'Missing required columns: {", ".join(missing)}'
            })
            return results

        # Process each row
        for row_num, row in enumerate(csv_reader, start=1):
            results['total_processed'] += 1

            try:
                # Check if faculty already exists (reuse your validation pattern)
                if Person.objects.filter(personid=row['employeeid']).exists():
                    results['skipped'] += 1
                    results['errors'].append({
                        'row': row_num,
                        'message': f'Employee ID {row["employeeid"]} already exists'
                    })
                    continue

                # Prepare data for FacultyForm (matching your exact field names)
                form_data = {
                    'employeeid': row['employeeid'],
                    'fname': row['fname'],
                    'lname': row['lname'],
                    'institutionalemail': row['institutionalemail'],
                    'personalemail': row.get('personalemail', ''),
                    'cnic': row.get('cnic', ''),
                    'gender': row.get('gender', 'M'),
                    'dob': parse_date(row.get('dob', '')),
                    'cnumber': row.get('cnumber', ''),
                    'designation': row['designation'],
                    'departmentid': get_department_by_id(row.get('departmentid', '')),
                    'joiningdate': parse_date(row.get('joiningdate', '')),
                    # Address fields
                    'country': row.get('country', 'Pakistan'),
                    'province': row.get('province', ''),
                    'city': row.get('city', ''),
                    'zipcode': parse_int(row.get('zipcode', '0')),
                    'streetaddress': row.get('streetaddress', ''),
                }

                # Create FacultyForm instance (reuse your form validation)
                form = FacultyForm(data=form_data)

                if form.is_valid():
                    # Convert CSV qualifications to POST-like format for your existing function
                    post_like_data = convert_csv_qualifications_to_post_format(row)

                    # Use your existing qualification extraction function
                    qualifications_data = student_views.extract_qualifications_from_post(post_like_data)

                    # Save faculty with qualifications (reuse your save method)
                    faculty = form.save_with_qualifications(qualifications_data, commit=True)
                    results['successful'] += 1
                else:
                    # Collect form errors
                    error_messages = []
                    for field, errors in form.errors.items():
                        error_messages.extend([f"{field}: {error}" for error in errors])

                    results['failed'] += 1
                    results['errors'].append({
                        'row': row_num,
                        'message': '; '.join(error_messages)
                    })

            except Exception as e:
                results['failed'] += 1
                results['errors'].append({
                    'row': row_num,
                    'message': f'Processing error: {str(e)}'
                })

        return results

    except Exception as e:
        results['errors'].append({
            'row': 0,
            'message': f'CSV parsing error: {str(e)}'
        })
        return results


# ===========================================
# CSV TEMPLATE DOWNLOADS
# ===========================================

@login_required
@user_passes_test(is_admin)
def download_csv_template(request, template_type):
    """Generate and download CSV templates - matches your exact form fields"""
    try:
        if template_type == 'students':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="student_import_template.csv"'

            writer = csv.writer(response)
            # Headers matching your StudentForm field names exactly
            writer.writerow([
                'studentid', 'fname', 'lname', 'institutionalemail', 'personalemail',
                'cnic', 'gender', 'dob', 'cnumber', 'programid', 'classid',
                'status', 'country', 'province', 'city', 'zipcode', 'streetaddress',
                # Dynamic qualification fields (1-3 qualifications)
                'qual1_degreetitle', 'qual1_educationboard', 'qual1_institution', 'qual1_passingyear',
                'qual1_totalmarks', 'qual1_obtainedmarks', 'qual1_iscurrent',
                'qual2_degreetitle', 'qual2_educationboard', 'qual2_institution', 'qual2_passingyear',
                'qual2_totalmarks', 'qual2_obtainedmarks', 'qual2_iscurrent',
                'qual3_degreetitle', 'qual3_educationboard', 'qual3_institution', 'qual3_passingyear',
                'qual3_totalmarks', 'qual3_obtainedmarks', 'qual3_iscurrent'
            ])

            # Sample row with your exact format requirements
            writer.writerow([
                'NUM-BSCS-2024-01', 'John', 'Doe', 'john.doe@university.edu', 'john@email.com',
                '12345-1234567-1', 'M', '2000-01-15', '+923001234567', '1', '1',
                'Enrolled', 'Pakistan', 'Punjab', 'Lahore', '54000', 'Sample Street',
                # Sample qualifications
                'Intermediate', 'BISE Lahore', 'Government College', '2018', '1100', '950', '0',
                'Matriculation', 'BISE Lahore', 'High School', '2016', '1050', '850', '0',
                '', '', '', '', '', '', ''  # Empty third qualification
            ])

        elif template_type == 'faculty':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="faculty_import_template.csv"'

            writer = csv.writer(response)
            # Headers matching your FacultyForm field names exactly
            writer.writerow([
                'employeeid', 'fname', 'lname', 'institutionalemail', 'personalemail',
                'cnic', 'gender', 'dob', 'cnumber', 'departmentid', 'designation',
                'joiningdate', 'country', 'province', 'city', 'zipcode', 'streetaddress',
                # Dynamic qualification fields (1-3 qualifications)
                'qual1_degreetitle', 'qual1_educationboard', 'qual1_institution', 'qual1_passingyear',
                'qual1_totalmarks', 'qual1_obtainedmarks', 'qual1_iscurrent',
                'qual2_degreetitle', 'qual2_educationboard', 'qual2_institution', 'qual2_passingyear',
                'qual2_totalmarks', 'qual2_obtainedmarks', 'qual2_iscurrent',
                'qual3_degreetitle', 'qual3_educationboard', 'qual3_institution', 'qual3_passingyear',
                'qual3_totalmarks', 'qual3_obtainedmarks', 'qual3_iscurrent'
            ])

            # Sample row with your exact format requirements
            writer.writerow([
                'FAC-2024-001', 'Jane', 'Smith', 'jane.smith@university.edu', 'jane@email.com',
                '12345-1234567-2', 'F', '1980-05-20', '+923001234568', '1', 'Assistant Professor',
                '2024-01-15', 'Pakistan', 'Punjab', 'Lahore', '54000', 'Sample Street',
                # Sample qualifications
                'PhD Computer Science', 'HEC', 'University of Punjab', '2015', '4.0', '3.8', '0',
                'MS Computer Science', 'HEC', 'University of Engineering', '2010', '4.0', '3.5', '0',
                'BS Computer Science', 'HEC', 'University of Technology', '2008', '4.0', '3.2', '0'
            ])

        else:
            return JsonResponse({'error': 'Invalid template type'}, status=400)

        return response

    except Exception as e:
        return JsonResponse({'error': f'Template generation failed: {str(e)}'}, status=500)


# ===========================================
# HELPER FUNCTIONS (Updated)
# ===========================================

def parse_date(date_string):
    """Parse date string with multiple format support"""
    if not date_string:
        return None

    formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']

    for fmt in formats:
        try:
            return datetime.strptime(date_string, fmt).date()
        except ValueError:
            continue

    return None


def parse_int(int_string):
    """Parse integer string with error handling"""
    if not int_string:
        return 0

    try:
        return int(int_string)
    except (ValueError, TypeError):
        return 0


def get_program_by_id(program_id):
    """Get program by ID with error handling"""
    if not program_id:
        return None

    try:
        return Program.objects.get(programid=program_id)
    except Program.DoesNotExist:
        return None


def get_class_by_id(class_id):
    """Get class by ID with error handling"""
    if not class_id:
        return None

    try:
        return Class.objects.get(classid=class_id)
    except Class.DoesNotExist:
        return None


def get_department_by_id(department_id):
    """Get department by ID with error handling"""
    if not department_id:
        return None

    try:
        return Department.objects.get(departmentid=department_id)
    except Department.DoesNotExist:
        return None


def convert_csv_qualifications_to_post_format(row):
    """
    Convert CSV qualification data to POST-like format that your existing function expects
    Your function expects: qualifications[INDEX][field] format
    """
    post_like_data = {}

    # Check for up to 3 qualifications in CSV
    qualification_index = 0

    for i in range(1, 4):  # qual1_, qual2_, qual3_
        prefix = f'qual{i}_'

        # Check if this qualification has data
        degree_title = row.get(f'{prefix}degreetitle', '').strip()

        if degree_title:  # Only process if degree title exists
            # Convert to your function's expected format: qualifications[INDEX][field]
            post_like_data[f'qualifications[{qualification_index}][degreetitle]'] = degree_title
            post_like_data[f'qualifications[{qualification_index}][educationboard]'] = row.get(
                f'{prefix}educationboard', '').strip()
            post_like_data[f'qualifications[{qualification_index}][institution]'] = row.get(f'{prefix}institution',
                                                                                            '').strip()
            post_like_data[f'qualifications[{qualification_index}][passingyear]'] = row.get(f'{prefix}passingyear',
                                                                                            '').strip()
            post_like_data[f'qualifications[{qualification_index}][totalmarks]'] = row.get(f'{prefix}totalmarks',
                                                                                           '').strip()
            post_like_data[f'qualifications[{qualification_index}][obtainedmarks]'] = row.get(f'{prefix}obtainedmarks',
                                                                                              '').strip()

            # Handle iscurrent boolean field
            iscurrent_value = row.get(f'{prefix}iscurrent', '').strip().lower()
            if iscurrent_value in ['1', 'true', 'yes']:
                post_like_data[f'qualifications[{qualification_index}][iscurrent]'] = '1'
            else:
                post_like_data[f'qualifications[{qualification_index}][iscurrent]'] = ''

            qualification_index += 1

    return post_like_data



# ===========================================
# DATA EXPORT FUNCTIONALITY
# ===========================================

@login_required
@user_passes_test(is_admin)
def export_data(request):
    """Export system data in various formats"""

    # Get statistics for the dashboard cards
    try:
        total_students = Student.objects.count()
        total_faculty = Faculty.objects.count()
        total_enrollments = Enrollment.objects.count()
        total_courses = Course.objects.count()
    except Exception as e:
        # Fallback in case of database issues
        total_students = 0
        total_faculty = 0
        total_enrollments = 0
        total_courses = 0
        messages.warning(request, f'Could not load statistics: {str(e)}')

    context = {
        'total_students': total_students,
        'total_faculty': total_faculty,
        'total_enrollments': total_enrollments,
        'total_courses': total_courses,
    }

    if request.method == 'POST':
        export_type = request.POST.get('export_type')
        data_type = request.POST.get('data_type')

        # Validate form data
        if not export_type or not data_type:
            messages.error(request, 'Please select both data type and export format')
            return render(request, 'admin/export_data.html', context)

        if export_type not in ['csv', 'excel']:
            messages.error(request, 'Invalid export format selected')
            return render(request, 'admin/export_data.html', context)

        if data_type not in ['students', 'faculty', 'enrollments']:
            messages.error(request, 'Invalid data type selected')
            return render(request, 'admin/export_data.html', context)

        try:
            # Get data based on type
            data, headers, rows = get_export_data(data_type)

            if not data.exists():
                messages.warning(request, f'No {data_type} data found to export')
                return render(request, 'admin/export_data.html', context)

            # Generate filename
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f'{data_type}_export_{timestamp}'

            # Handle different export types
            if export_type == 'csv':
                return generate_csv_export(data_type, headers, rows, filename)
            elif export_type == 'excel':
                return generate_excel_export(data_type, headers, rows, filename)

        except Exception as e:
            messages.error(request, f'Export failed: {str(e)}')
            return render(request, 'admin/export_data.html', context)

    return render(request, 'admin/export_data.html', context)


def get_export_data(data_type):
    """Get data, headers, and rows for export based on data type"""

    if data_type == 'students':
        data = Student.objects.select_related(
            'studentid',
            'studentid__address',
            'programid',
            'programid__departmentid',
            'studentid__qualification',
        ).all()

        headers = [
            'Student ID', 'First Name', 'Last Name','Gender','Date of birth','CNIC' ,'Institutional Email','Personal Email',
            'Phone', 'Program', 'Department', 'Batch Year', 'Status', 'Province','City','Zip Code','Street Address',
        ]

        rows = []
        for student in data:
            rows.append([
                getattr(student.studentid, 'personid', '') if student.studentid else '',
                getattr(student.studentid, 'fname', '') if student.studentid else '',
                getattr(student.studentid, 'lname', '') if student.studentid else '',
                getattr(student.studentid, 'gender', '') if student.studentid else '',
                getattr(student.studentid, 'dob','') if student.studentid else '',
                getattr(student.studentid, 'cnic','') if student.studentid else '',
                getattr(student.studentid, 'institutionalemail', '') if student.studentid else '',
                getattr(student.studentid, 'personalemail', '') if student.studentid else '',
                getattr(student.studentid, 'cnumber', '') if student.studentid else '',
                getattr(student.programid, 'programname', '') if student.programid else '',
                getattr(student.programid.departmentid, 'departmentname',
                        '') if student.programid and student.programid.departmentid else '',
                getattr(student.classid, 'batchyear', '') if hasattr(student.classid, 'batchyear') else '',
                getattr(student,'status','') if hasattr(student,'status') else '',
                getattr(student.studentid.address,'province','') if hasattr (student.studentid.address, 'province') else '',
                getattr(student.studentid.address, 'city', '') if hasattr(student.studentid.address,'city') else '',
                getattr(student.studentid.address, 'zipcode','') if hasattr(student.studentid.address, 'zipcode') else '',
                getattr(student.studentid.address, 'streetaddress','') if hasattr(student.studentid.address, 'streetaddress') else '',

            ])

    elif data_type == 'faculty':
        data = Faculty.objects.select_related(
            'employeeid',
            'departmentid'
        ).all()

        headers = [
            'Employee ID', 'First Name', 'Last Name', 'Email',
            'Phone', 'Department', 'Designation', 'Qualification',
            'Join Date', 'Status'
        ]

        rows = []
        for faculty in data:
            rows.append([
                getattr(faculty.employeeid, 'employeeid', '') if faculty.employeeid else '',
                getattr(faculty.employeeid, 'firstname', '') if faculty.employeeid else '',
                getattr(faculty.employeeid, 'lastname', '') if faculty.employeeid else '',
                getattr(faculty.employeeid, 'email', '') if faculty.employeeid else '',
                getattr(faculty.employeeid, 'phone', '') if faculty.employeeid else '',
                getattr(faculty.departmentid, 'departmentname', '') if faculty.departmentid else '',
                getattr(faculty, 'designation', '') if hasattr(faculty, 'designation') else '',
                getattr(faculty, 'qualification', '') if hasattr(faculty, 'qualification') else '',
                getattr(faculty.employeeid, 'date_joined', '').strftime('%Y-%m-%d') if faculty.employeeid and hasattr(
                    faculty.employeeid, 'date_joined') and faculty.employeeid.date_joined else '',
                'Active'  # You can customize this based on your faculty model
            ])

    elif data_type == 'enrollments':
        data = Enrollment.objects.select_related(
            'studentid__studentid',
            'allocationid__coursecode',
            'allocationid__sessionid',
            'allocationid__classid'
        ).all()

        headers = [
            'Student ID', 'Student Name', 'Course Code', 'Course Name',
            'Session', 'Class', 'Enrollment Date', 'Status'
        ]

        rows = []
        for enrollment in data:
            student_name = ''
            if enrollment.studentid and enrollment.studentid.studentid:
                firstname = getattr(enrollment.studentid.studentid, 'firstname', '') or ''
                lastname = getattr(enrollment.studentid.studentid, 'lastname', '') or ''
                student_name = f"{firstname} {lastname}".strip()

            rows.append([
                getattr(enrollment.studentid.studentid, 'studentid',
                        '') if enrollment.studentid and enrollment.studentid.studentid else '',
                student_name,
                getattr(enrollment.allocationid.coursecode, 'coursecode',
                        '') if enrollment.allocationid and enrollment.allocationid.coursecode else '',
                getattr(enrollment.allocationid.coursecode, 'coursename',
                        '') if enrollment.allocationid and enrollment.allocationid.coursecode else '',
                getattr(enrollment.allocationid.sessionid, 'sessionname',
                        '') if enrollment.allocationid and enrollment.allocationid.sessionid else '',
                getattr(enrollment.allocationid.classid, 'classname',
                        '') if enrollment.allocationid and enrollment.allocationid.classid else '',
                getattr(enrollment, 'enrollmentdate', '').strftime('%Y-%m-%d') if hasattr(enrollment,
                                                                                          'enrollmentdate') and enrollment.enrollmentdate else '',
                'Enrolled'  # You can customize this based on your enrollment model
            ])

    return data, headers, rows


def generate_csv_export(data_type, headers, rows, filename):
    """Generate CSV export response"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    response['Content-Type'] = 'text/csv; charset=utf-8'

    # Add BOM for proper UTF-8 encoding in Excel
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)

    return response


def generate_excel_export(data_type, headers, rows, filename):
    """Generate Excel export response"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = data_type.capitalize()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add headers
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}1']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style

        # Auto-adjust column width
        ws.column_dimensions[col_letter].width = max(len(header) + 2, 12)

    # Add data rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{row_num}']
            cell.value = cell_value
            cell.border = border_style

            # Adjust column width based on content
            current_width = ws.column_dimensions[col_letter].width
            content_width = len(str(cell_value)) + 2
            if content_width > current_width:
                ws.column_dimensions[col_letter].width = min(content_width, 50)  # Max width of 50

    # Create a table for better formatting
    if rows:
        table_range = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        table = Table(displayName=f"{data_type}Table", ref=table_range)

        # Add a table style
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Add metadata
    ws['A' + str(len(rows) + 3)] = f"Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A' + str(len(rows) + 4)] = f"Total Records: {len(rows)}"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    wb.save(response)
    return response





